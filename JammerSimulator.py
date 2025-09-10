"""JammerSimulator - Núcleo GUI simplificado

Este archivo contiene únicamente:
	- Carga de parámetros
	- Clases básicas (Satellite, Constellation, LEOEducationalCalculations)
	- Núcleo mínimo (JammerSimulatorCore) con parámetros usados por la GUI
	- Interfaz Tkinter con animación orbital y métricas

Se ha eliminado todo el código de demostración por consola / selftest para
mantener el script compacto y orientado a despliegue GUI.
"""
from __future__ import annotations

import json
import math
import time
import csv
from dataclasses import dataclass
from typing import List, Optional, Dict, Any

# Importar sistema de jammers
try:
	from JammerSystem import JammerManager
	JAMMERS_AVAILABLE = True
except ImportError as e:
	print(f"Warning: JammerSystem no disponible: {e}")
	JammerManager = None
	JAMMERS_AVAILABLE = False

# Intentamos importar tkinter al cargar el módulo; si falla, se gestiona en main().
try:
	import tkinter as tk
	from tkinter import ttk
except ImportError:  # entorno sin GUI
	tk = None
	ttk = None

SPEED_OF_LIGHT = 299_792_458.0  # m/s
EARTH_RADIUS_M = 6_371_000.0
MU_EARTH = 3.986004418e14  # m^3/s^2 (fase 1)

# Constantes de rotación terrestre para animación realista
EARTH_ROTATION_PERIOD_S = 24 * 3600  # 86400 segundos (día sidéreo real ~86164s, pero usamos día solar)
EARTH_ROTATION_DEG_PER_S = 360.0 / EARTH_ROTATION_PERIOD_S  # ~0.004167 deg/s

# Tipos para enlaces separados
LinkSense = str  # 'UL' o 'DL'

@dataclass
class LinkInputs:
	f_Hz: float
	B_Hz: float 
	EIRP_dBW: float
	GT_dBK: float
	L_extra_dB: float

@dataclass 
class LinkOutputs:
	FSPL_dB: float
	CN0_dBHz: float
	CN_dB: float
	visible: bool
	latency_ms: float

# ------------------------------------------------------------- #
# Funciones puras para cálculos de enlace                       #
# ------------------------------------------------------------- #
def fspl_dB(f_Hz: float, d_m: float) -> float:
	"""Free Space Path Loss en dB."""
	if d_m <= 0 or f_Hz <= 0:
		return float('inf')
	return 20 * math.log10(4 * math.pi * d_m * f_Hz / SPEED_OF_LIGHT)

def cn0_dBHz(EIRP_dBW: float, GT_dBK: float, FSPL_dB: float, L_extra_dB: float) -> float:
	"""C/N0 en dBHz usando la fórmula estándar."""
	return EIRP_dBW + GT_dBK - FSPL_dB - L_extra_dB + 228.6

def cn_dB(CN0_dBHz: float, B_Hz: float) -> float:
	"""C/N en dB a partir de C/N0 y ancho de banda."""
	if B_Hz <= 0:
		return float('-inf')
	return CN0_dBHz - 10 * math.log10(B_Hz)

def combine_end_to_end(cnUL_dB: float, cnDL_dB: float) -> dict:
	"""Combina C/N de UL y DL usando suma lineal de (N/C)."""
	if math.isnan(cnUL_dB) or math.isnan(cnDL_dB) or math.isinf(cnUL_dB) or math.isinf(cnDL_dB):
		return {"NC_tot_dB": float('nan'), "CN_tot_dB": float('nan'), "CINR_tot_dB": float('nan')}
	
	# Convertir C/N a N/C en escala lineal
	nc_ul = 10.0 ** (-cnUL_dB / 10.0)
	nc_dl = 10.0 ** (-cnDL_dB / 10.0)
	nc_tot = nc_ul + nc_dl
	
	# Convertir de vuelta a dB
	if nc_tot <= 0:
		return {"NC_tot_dB": float('inf'), "CN_tot_dB": float('-inf'), "CINR_tot_dB": float('-inf')}
	
	nc_tot_dB = 10 * math.log10(nc_tot)
	cn_tot_dB = -nc_tot_dB
	cinr_tot_dB = cn_tot_dB  # Por ahora, sin interferencia
	
	return {"NC_tot_dB": nc_tot_dB, "CN_tot_dB": cn_tot_dB, "CINR_tot_dB": cinr_tot_dB}

# ------------------------------------------------------------- #
# Helpers dB / lineal (Fase 0)                                  #
# ------------------------------------------------------------- #
def lin_to_db(x: float, min_lin: float = 1e-30) -> float:
	"""10log10(x) protegido contra valores <=0."""
	if x <= min_lin:
		x = min_lin
	return 10.0 * math.log10(x)


def db_to_lin(x_db: float) -> float:
	return 10.0 ** (x_db / 10.0)


def sum_powers_db(terms_db: List[float]) -> float:
	"""Suma de potencias en dominio dB (ignora NaN)."""
	linear_sum = 0.0
	for v in terms_db:
		if v is None or (isinstance(v, float) and math.isnan(v)):
			continue
		linear_sum += db_to_lin(v)
	if linear_sum <= 0:
		return float('-inf')
	return lin_to_db(linear_sum)

# ------------------------------------------------------------- #
# Carga de parámetros centralizada                              #
# ------------------------------------------------------------- #
class ParameterLoader:
	def __init__(self, filename: str = "SimulatorParameters.json"):
		with open(filename, 'r', encoding='utf-8') as f:
			self.data = json.load(f)

	def get(self, path: List[str]):
		ref = self.data
		for p in path:
			ref = ref[p]
		return ref

# ------------------------------------------------------------- #
# Modelo básico de un satélite (ahora solo necesitamos altitud) #
# ------------------------------------------------------------- #
@dataclass
class Satellite:
	name: str
	altitude_m: float  # Altitud media sobre superficie
	orbital_angle_deg: float = 0.0  # Posición angular en la órbita
	constellation_id: str = "default"  # ID de la constelación a la que pertenece


class Constellation:
	"""Representa una colección de satélites en una órbita específica.
	
	Preparado para futuras expansiones:
	- Múltiples constelaciones LEO/GEO
	- Diferentes planos orbitales
	- Handovers entre satélites
	- Propagación temporal coordinada
	"""
	def __init__(self, satellites: List[Satellite], constellation_type: str = "LEO", constellation_id: str = "default"):
		self.satellites = satellites
		self.constellation_type = constellation_type  # "LEO" o "GEO"
		self.constellation_id = constellation_id
		self.orbital_parameters = {}  # Para futuros parámetros específicos

	@classmethod
	def single_leo(cls, altitude_m: float, constellation_id: str = "LEO-1") -> 'Constellation':
		satellite = Satellite(name=f"{constellation_id}-SAT-1", altitude_m=altitude_m, constellation_id=constellation_id)
		return cls([satellite], "LEO", constellation_id)
	
	@classmethod
	def single_geo(cls, altitude_m: float = 35_786_000.0, constellation_id: str = "GEO-1") -> 'Constellation':
		satellite = Satellite(name=f"{constellation_id}-SAT-1", altitude_m=altitude_m, constellation_id=constellation_id)
		return cls([satellite], "GEO", constellation_id)
	
	def get_active_satellite(self) -> Optional[Satellite]:
		"""Retorna el satélite activo (para simulación simple, es el primero)."""
		return self.satellites[0] if self.satellites else None
	
	def add_satellite(self, satellite: Satellite):
		"""Añade un satélite a la constelación."""
		satellite.constellation_id = self.constellation_id
		self.satellites.append(satellite)
	
	def get_satellites_by_visibility(self, observer_position: tuple = (0, 0)) -> List[Satellite]:
		"""Futura función para filtrar satélites visibles desde una posición."""
		# Placeholder para futura implementación de visibilidad múltiple
		return self.satellites


class MultiConstellation:
	"""Gestor de múltiples constelaciones para futuras expansiones del simulador.
	
	Permitirá manejar:
	- Múltiples constelaciones LEO con diferentes altitudes
	- Constelaciones GEO con diferentes posiciones
	- Handovers automáticos entre satélites
	- Análisis de coverage combinado
	"""
	def __init__(self):
		self.constellations: Dict[str, Constellation] = {}
		self.active_constellation_id: Optional[str] = None
	
	def add_constellation(self, constellation: Constellation):
		"""Añade una constelación al sistema."""
		self.constellations[constellation.constellation_id] = constellation
		if self.active_constellation_id is None:
			self.active_constellation_id = constellation.constellation_id
	
	def get_active_constellation(self) -> Optional[Constellation]:
		"""Retorna la constelación actualmente activa."""
		if self.active_constellation_id:
			return self.constellations.get(self.active_constellation_id)
		return None
	
	def set_active_constellation(self, constellation_id: str):
		"""Cambia la constelación activa."""
		if constellation_id in self.constellations:
			self.active_constellation_id = constellation_id
	
	def get_all_visible_satellites(self, observer_position: tuple = (0, 0)) -> List[tuple]:
		"""Retorna todos los satélites visibles de todas las constelaciones.
		Retorna lista de tuplas (satellite, constellation_id)."""
		visible_satellites = []
		for const_id, constellation in self.constellations.items():
			for satellite in constellation.get_satellites_by_visibility(observer_position):
				visible_satellites.append((satellite, const_id))
		return visible_satellites
	
	def get_best_satellite_for_handover(self) -> Optional[tuple]:
		"""Futura función para seleccionar el mejor satélite para handover."""
		# Placeholder para lógica de handover inteligente
		return None


# ------------------------------------------------------------- #
# Calculos para los Enlaces #
# ------------------------------------------------------------- #
class LEOEducationalCalculations:
	def __init__(self, altitude_m: float, frequency_hz: float = 12e9):
		self.altitude_m = altitude_m
		self.frequency_hz = frequency_hz
		# Parámetros de potencia/noise por ahora sencillos; se podrán ajustar
		self.default_bandwidth_hz = 1e6  # 1 MHz para simplificar (coherente con Paso 2)
		# Constantes para ecuación educativa C/N0
		self.K_BOLTZ_TERM = 228.6  # dB (10log10(1/k))

	def slant_range_simple(self, elevation_deg: float) -> float:
		"""Distancia aproximada (m) usando d ≈ h / sin(E).

		Elección consciente: aproximación para crear intuición. Más adelante
		se añadirá la fórmula exacta con radio terrestre.
		"""
		if elevation_deg <= 0:
			elevation_deg = 0.1  # evitar división por cero manteniendo significado
		elev_rad = math.radians(elevation_deg)
		return self.altitude_m / math.sin(elev_rad)

	def free_space_path_loss_db(self, distance_m: float) -> float:
		return 20 * math.log10(4 * math.pi * distance_m * self.frequency_hz / SPEED_OF_LIGHT)

	# Latencia de propagación (one-way o round trip)
	def propagation_delay_ms(self, distance_m: float, round_trip: bool = False) -> float:
		factor = 2.0 if round_trip else 1.0
		return (factor * distance_m / SPEED_OF_LIGHT) * 1000.0

	# C/N0 simplificado
	def cn0_dbhz(self, eirp_dbw: float, gt_dbk: float, fspl_db: float) -> float:
		return eirp_dbw + gt_dbk - fspl_db + self.K_BOLTZ_TERM

	# C/N a partir de C/N0 y ancho de banda
	def cn_db(self, cn0_dbhz: float, bandwidth_hz: float | None = None) -> float:
		bw = bandwidth_hz or self.default_bandwidth_hz
		return cn0_dbhz - 10 * math.log10(bw)
class JammerSimulatorCore:
	"""Núcleo mínimo usado por la GUI (parámetros y cálculos)."""

	def __init__(self, params: ParameterLoader):
		self.params = params
		leo_alt = params.get(["LEO", "Altitude"])
		self.constellation = Constellation.single_leo(altitude_m=leo_alt)
		self.calc = LEOEducationalCalculations(altitude_m=leo_alt)
		# Parámetros de potencia base
		self.eirp_dbw = params.get(["LEO", "EIRP", "base"])  # Ejemplo
		self.gt_dbk = params.get(["LEO", "G_T", "base"])     # Placeholder
		# GEO (si disponible)
		try:
			self.geo_altitude_m = params.get(["GEO", "Altitude"])
			self.geo_eirp_dbw = params.get(["GEO", "EIRP", "base"])
			self.geo_gt_dbk = params.get(["GEO", "G_T", "base"])
		except Exception:
			# Valores típicos por defecto
			self.geo_altitude_m = 35_786_000.0
			self.geo_eirp_dbw = 52.0
			self.geo_gt_dbk = -5.0
		# Calculadora separada para GEO (misma frecuencia inicial)
		self.geo_calc = LEOEducationalCalculations(altitude_m=self.geo_altitude_m, frequency_hz=self.calc.frequency_hz)
		# --------------------------------------------------------- #
		# Contenedores estructurados (Fase 0)                      #
		# --------------------------------------------------------- #
		self.losses: Dict[str, float] = {
			'RFL_feeder': 0.0,
			'AML_misalignment': 0.0,
			'AA_atmos': 0.0,
			'Rain_att': 0.0,
			'PL_polarization': 0.0,
			'L_pointing': 0.0,
			'L_impl': 0.0,
		}
		self.noise: Dict[str, float] = {
			'T_rx': 120.0,
			'T_clear_sky': 30.0,
			'T_rain_excess': 0.0,
		}
		self.power: Dict[str, float] = {
			'EIRP_sat_saturated': self.eirp_dbw,
			'Input_backoff': 0.0,
		}
		self.throughput: Dict[str, float] = {
			'Rb_Mbps': 10.0,
			'EbN0_req_dB': 4.0,
		}
		self.latencies: Dict[str, float] = {
			'Processing_delay_ms': 2.0,
			'Switching_delay_ms': 1.0,
		}
		self.coverage: Dict[str, float] = {
			'Min_elevation_deg': 10.0,
		}
		
		# Parámetros de enlaces separados por constelación
		try:
			# Intentar cargar enlaces LEO específicos (por defecto LEO)
			self.link_presets = {
				'LEO': {
					'UL': params.get(["LEO", "Links", "UL"]),
					'DL': params.get(["LEO", "Links", "DL"])
				},
				'GEO': {
					'UL': params.get(["GEO", "Links", "UL"]),
					'DL': params.get(["GEO", "Links", "DL"])
				}
			}
		except Exception:
			# Fallback si no existe la estructura LEO/GEO separada
			try:
				# Intentar estructura antigua global
				self.link_presets = {
					'LEO': {
						'UL': params.get(["Links", "UL"]),
						'DL': params.get(["Links", "DL"])
					},
					'GEO': {
						'UL': params.get(["Links", "UL"]),
						'DL': params.get(["Links", "DL"])
					}
				}
			except Exception:
				# Fallback hardcoded
				self.link_presets = {
					'LEO': {
						'UL': {"freq_GHz": 30.0, "bw_MHz": 20.0, "EIRP_dBW": 45.0, "GT_dBK": -8.0, "extra_losses_dB": 0.0},
						'DL': {"freq_GHz": 20.0, "bw_MHz": 20.0, "EIRP_dBW": 48.0, "GT_dBK": 12.0, "extra_losses_dB": 0.0}
					},
					'GEO': {
						'UL': {"freq_GHz": 14.0, "bw_MHz": 12.0, "EIRP_dBW": 62.0, "GT_dBK": -2.0, "extra_losses_dB": 0.0},
						'DL': {"freq_GHz": 11.7, "bw_MHz": 12.0, "EIRP_dBW": 56.0, "GT_dBK": 8.0, "extra_losses_dB": 0.0}
					}
				}
	
	def compute_link_outputs(self, inputs: LinkInputs, d_km: float) -> LinkOutputs:
		"""Calcula las métricas de salida para un sentido de enlace."""
		# Verificar visibilidad basada en elevación y distancia válida
		elevation_deg = getattr(self, 'current_elevation_deg', 0.0)
		visible = d_km > 0 and elevation_deg > 0
		
		if not visible:
			return LinkOutputs(
				FSPL_dB=float('nan'),
				CN0_dBHz=float('nan'), 
				CN_dB=float('nan'),
				visible=False,
				latency_ms=float('nan')
			)
		
		# Cálculos usando las funciones puras
		fspl = fspl_dB(inputs.f_Hz, d_km * 1000)  # convertir km a m
		cn0 = cn0_dBHz(inputs.EIRP_dBW, inputs.GT_dBK, fspl, inputs.L_extra_dB)
		cn = cn_dB(cn0, inputs.B_Hz)
		latency = (d_km * 1000 / SPEED_OF_LIGHT) * 1000  # ms
		
		return LinkOutputs(
			FSPL_dB=fspl,
			CN0_dBHz=cn0,
			CN_dB=cn,
			visible=True,
			latency_ms=latency
		)


# ------------------------------------------------------------- #
# GUI (se define a nivel de módulo; solo se usará si tkinter disponible)
# ------------------------------------------------------------- #
class SimulatorGUI:
	def __init__(self, root, core: JammerSimulatorCore):
		self.root = root
		self.core = core
		self.root.title("Jammer Simulator (LEO/GEO)")
		self.running = False
		self.mode_var = None
		
		# Variables de rotación y posición orbital
		self.orbit_angle_deg = 0.0
		self.step_orbit_deg = 0.2  # Reducido para mayor granularidad en elevación
		self.earth_rotation_angle_deg = 0.0  # Rotación acumulada de la Tierra
		self.geo_rotation_angle_deg = 0.0    # Rotación GEO (sincronizada con Tierra)
		
		# Estado de enlaces separados (iniciará con LEO por defecto)
		self.link_state = {
			'UL': LinkInputs(0, 0, 0, 0, 0),
			'DL': LinkInputs(0, 0, 0, 0, 0)
		}
		self.link_out = {
			'UL': LinkOutputs(0.0, 0.0, 0.0, False, 0.0),
			'DL': LinkOutputs(0.0, 0.0, 0.0, False, 0.0)
		}
		self.current_link_sense = 'UL'  # Pestaña activa
		
		# Configuración orbital
		alt_km = self.core.constellation.satellites[0].altitude_m / 1000.0
		self.Re_km = EARTH_RADIUS_M / 1000.0
		self.orbit_r_km = self.Re_km + alt_km
		self.geo_orbit_r_km = self.Re_km + self.core.geo_altitude_m/1000.0
		self.horizon_central_angle_deg = math.degrees(math.acos(self.Re_km / self.orbit_r_km))
		self.orbit_angle_deg = (360.0 - self.horizon_central_angle_deg)
		
		# Parámetros de animación desde JSON
		try:
			self.time_scale_factor = self.core.params.get(["Animation", "time_scale_factor"])
			self.animation_interval_ms = int(self.core.params.get(["Animation", "update_interval_ms"]))
		except:
			# Valores por defecto si no están en JSON
			self.time_scale_factor = 1000.0
			self.animation_interval_ms = 300
		
		# Datos de simulación
		self.history: List[Dict[str, Any]] = []
		self.start_time: Optional[float] = None
		self.last_animation_time: Optional[float] = None
		
		# Control de tiempo manual vs automático
		self.manual_time_control = False  # Nuevo: modo manual vs automático
		self.simulation_time_s = 0.0      # Tiempo de simulación actual
		self.max_simulation_time_s = 7200.0  # Máximo 2 horas de simulación (más órbitas LEO completas)
		
		# Inicializar variables de estado orbital
		self.current_elevation_deg = 0.0
		self.current_visible = False
		self.current_slant_distance_m = 0.0
		
		# Inicializar gestor de jammers
		self.jammer_manager = None  # Se inicializará en _build_layout
		
		self._build_layout(); self._draw_static(); self._load_link_presets_for_mode(); self._refresh_gui_values(); self.update_metrics()
	
	def _refresh_gui_values(self):
		"""Actualiza los valores en la GUI después de la inicialización."""
		if hasattr(self, 'eirp_var') and hasattr(self, 'gt_var') and hasattr(self, 'bw_var'):
			# Asegurar que los valores básicos estén establecidos
			if self.eirp_var.get() == 0.0:
				self.eirp_var.set(self.core.eirp_dbw)
			if self.gt_var.get() == 0.0:
				self.gt_var.set(self.core.gt_dbk)
			if self.bw_var.get() == 0.0:
				self.bw_var.set(self.core.calc.default_bandwidth_hz/1e6)

	def _load_link_presets_for_mode(self):
		"""Carga los parámetros UL/DL según el modo activo (LEO/GEO)."""
		current_mode = getattr(self, 'mode_var', None)
		mode = current_mode.get() if current_mode else 'LEO'
		
		# Cargar presets según modo
		try:
			ul_preset = self.core.link_presets[mode]['UL']
			dl_preset = self.core.link_presets[mode]['DL']
		except KeyError:
			# Fallback a LEO si el modo no existe
			ul_preset = self.core.link_presets['LEO']['UL']
			dl_preset = self.core.link_presets['LEO']['DL']
		
		# Actualizar estado de enlaces
		self.link_state['UL'] = LinkInputs(
			f_Hz=ul_preset['freq_GHz'] * 1e9,
			B_Hz=ul_preset['bw_MHz'] * 1e6,
			EIRP_dBW=ul_preset['EIRP_dBW'],
			GT_dBK=ul_preset['GT_dBK'],
			L_extra_dB=ul_preset['extra_losses_dB']
		)
		self.link_state['DL'] = LinkInputs(
			f_Hz=dl_preset['freq_GHz'] * 1e9,
			B_Hz=dl_preset['bw_MHz'] * 1e6,
			EIRP_dBW=dl_preset['EIRP_dBW'],
			GT_dBK=dl_preset['GT_dBK'],
			L_extra_dB=dl_preset['extra_losses_dB']
		)
		
		# Actualizar variables GUI si existen
		if hasattr(self, 'ul_freq_var'):
			self.ul_freq_var.set(ul_preset['freq_GHz'])
			self.ul_bw_var.set(ul_preset['bw_MHz'])
			self.ul_eirp_var.set(ul_preset['EIRP_dBW'])
			self.ul_gt_var.set(ul_preset['GT_dBK'])
			self.ul_losses_var.set(ul_preset['extra_losses_dB'])
		
		if hasattr(self, 'dl_freq_var'):
			self.dl_freq_var.set(dl_preset['freq_GHz'])
			self.dl_bw_var.set(dl_preset['bw_MHz'])
			self.dl_eirp_var.set(dl_preset['EIRP_dBW'])
			self.dl_gt_var.set(dl_preset['GT_dBK'])
			self.dl_losses_var.set(dl_preset['extra_losses_dB'])

	def _build_layout(self):
		self.mainframe = ttk.Frame(self.root, padding=5); self.mainframe.pack(fill='both', expand=True)
		
		# Crear frame izquierdo con scroll
		left_container = ttk.Frame(self.mainframe)
		left_container.pack(side='left', fill='y')
		
		# Canvas y scrollbar para el contenido izquierdo
		self.left_canvas = tk.Canvas(left_container, width=350, highlightthickness=0)
		left_scrollbar = ttk.Scrollbar(left_container, orient='vertical', command=self.left_canvas.yview)
		self.scrollable_left = ttk.Frame(self.left_canvas)
		
		self.scrollable_left.bind(
			"<Configure>",
			lambda e: self.left_canvas.configure(scrollregion=self.left_canvas.bbox("all"))
		)
		
		self.left_canvas.create_window((0, 0), window=self.scrollable_left, anchor="nw")
		self.left_canvas.configure(yscrollcommand=left_scrollbar.set)
		
		self.left_canvas.pack(side='left', fill='both', expand=True)
		left_scrollbar.pack(side='right', fill='y')
		
		# Soporte para scroll con rueda del ratón en la columna izquierda
		def _on_left_mousewheel(event):
			if hasattr(self, 'left_canvas'):
				delta = int(-1*(event.delta/120))
				self.left_canvas.yview_scroll(delta, 'units')
		
		self.left_canvas.bind('<MouseWheel>', _on_left_mousewheel)
		self.scrollable_left.bind('<MouseWheel>', _on_left_mousewheel)
		
		# También bind para cuando se entra/sale del área
		def _bind_to_mousewheel(event):
			self.left_canvas.bind_all('<MouseWheel>', _on_left_mousewheel)
		def _unbind_from_mousewheel(event):
			self.left_canvas.unbind_all('<MouseWheel>')
		
		self.left_canvas.bind('<Enter>', _bind_to_mousewheel)
		self.left_canvas.bind('<Leave>', _unbind_from_mousewheel)
		
		# Ahora usar scrollable_left en lugar de left para todos los controles
		left = self.scrollable_left
		
		bold_lbl = ('Segoe UI', 10, 'bold')
		
		# -------- Selector de Modo --------
		ttk.Label(left, text="Modo:", font=bold_lbl).pack(anchor='w')
		self.mode_var = tk.StringVar(value='LEO')
		mode_combo = ttk.Combobox(left, textvariable=self.mode_var, values=['LEO','GEO'], state='readonly')
		mode_combo.pack(anchor='w', pady=2)
		mode_combo.bind('<<ComboboxSelected>>', lambda e: self._change_mode())
		
		# -------- Parámetros Básicos --------
		basic_frame = ttk.LabelFrame(left, text='Parámetros Básicos')
		basic_frame.pack(fill='x', pady=6)
		
		ttk.Label(basic_frame, text="EIRP (dBW):", font=bold_lbl).pack(anchor='w', padx=5, pady=1)
		self.eirp_var = tk.DoubleVar(value=self.core.eirp_dbw)
		ttk.Entry(basic_frame, textvariable=self.eirp_var, width=10).pack(anchor='w', padx=5)
		
		ttk.Label(basic_frame, text="G/T (dB/K):", font=bold_lbl).pack(anchor='w', padx=5, pady=1)
		self.gt_var = tk.DoubleVar(value=self.core.gt_dbk)
		ttk.Entry(basic_frame, textvariable=self.gt_var, width=10).pack(anchor='w', padx=5)
		
		ttk.Label(basic_frame, text="BW (MHz):", font=bold_lbl).pack(anchor='w', padx=5, pady=1)
		self.bw_var = tk.DoubleVar(value=self.core.calc.default_bandwidth_hz/1e6)
		ttk.Entry(basic_frame, textvariable=self.bw_var, width=10).pack(anchor='w', padx=5, pady=(0,5))
		
		# -------- Panel de Jammers Adaptativo (Escenario 2) --------
		if JAMMERS_AVAILABLE and JammerManager:
			self.jammer_manager = JammerManager(left)
			jammer_panel = self.jammer_manager.get_panel()
			jammer_panel.pack(fill='x', pady=6)
		else:
			self.jammer_manager = None
		
		# -------- Pestañas de Enlaces UL/DL/End-to-End --------
		links_frame = ttk.LabelFrame(left, text='Enlaces Separados')
		links_frame.pack(fill='x', pady=6)
		
		self.notebook = ttk.Notebook(links_frame)
		self.notebook.pack(fill='both', expand=True, padx=2, pady=2)
		
		# Crear pestañas
		self._create_uplink_tab()
		self._create_downlink_tab() 
		self._create_endtoend_tab()
		
		# Bind para cambio de pestaña
		self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)
		
		# -------- Potencia / Backoff (Fase 3) --------
		power_frame = ttk.LabelFrame(left, text='Potencia / Backoff')
		power_frame.pack(fill='x', pady=6)
		self.eirp_sat_var = tk.DoubleVar(value=self.core.power['EIRP_sat_saturated'])
		self.input_bo_var = tk.DoubleVar(value=self.core.power['Input_backoff'])
		self.manual_override_var = tk.BooleanVar(value=False)
		self.manual_eirp_var = tk.DoubleVar(value=self.core.eirp_dbw)
		row_pf1 = ttk.Frame(power_frame); row_pf1.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_pf1, text='EIRP Saturado (dBW):', font=bold_lbl).pack(side='left')
		self.eirp_sat_entry = ttk.Entry(row_pf1, textvariable=self.eirp_sat_var, width=8); self.eirp_sat_entry.pack(side='right')
		row_pf2 = ttk.Frame(power_frame); row_pf2.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_pf2, text='Back-off Entrada (dB):', font=bold_lbl).pack(side='left')
		self.input_bo_entry = ttk.Entry(row_pf2, textvariable=self.input_bo_var, width=8); self.input_bo_entry.pack(side='right')
		row_pf3 = ttk.Frame(power_frame); row_pf3.pack(fill='x', padx=2, pady=1)
		self.output_bo_label = ttk.Label(row_pf3, text='Back-off Salida: 0.0 dB', font=bold_lbl)
		self.output_bo_label.pack(side='left', anchor='w')
		self.eirp_eff_label = ttk.Label(row_pf3, text='EIRP Efectivo: -- dBW', font=bold_lbl)
		self.eirp_eff_label.pack(side='right', anchor='e')
		row_pf4 = ttk.Frame(power_frame); row_pf4.pack(fill='x', padx=2, pady=1)
		self.override_chk = ttk.Checkbutton(row_pf4, text='Override EIRP', variable=self.manual_override_var, command=lambda: self.update_metrics())
		self.override_chk.pack(side='left')
		self.manual_eirp_entry = ttk.Entry(row_pf4, textvariable=self.manual_eirp_var, width=8)
		self.manual_eirp_entry.pack(side='right')
		sep2 = ttk.Separator(left, orient='horizontal'); sep2.pack(fill='x', pady=6)
		# -------- Ruido (Entradas) --------
		ruido_frame = ttk.LabelFrame(left, text='Ruido (Entradas)')
		ruido_frame.pack(fill='x', pady=6)
		self.t_rx_var = tk.DoubleVar(value=self.core.noise['T_rx'])
		self.t_sky_var = tk.DoubleVar(value=self.core.noise['T_clear_sky'])
		self.t_rain_var = tk.DoubleVar(value=self.core.noise['T_rain_excess'])
		row_n1 = ttk.Frame(ruido_frame); row_n1.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_n1, text='T RX (K):', font=bold_lbl).pack(side='left')
		self.t_rx_entry = ttk.Entry(row_n1, textvariable=self.t_rx_var, width=7); self.t_rx_entry.pack(side='right')
		row_n2 = ttk.Frame(ruido_frame); row_n2.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_n2, text='T Cielo (K):', font=bold_lbl).pack(side='left')
		self.t_sky_entry = ttk.Entry(row_n2, textvariable=self.t_sky_var, width=7); self.t_sky_entry.pack(side='right')
		row_n3 = ttk.Frame(ruido_frame); row_n3.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_n3, text='T Exceso Lluvia (K):', font=bold_lbl).pack(side='left')
		self.t_rain_entry = ttk.Entry(row_n3, textvariable=self.t_rain_var, width=7); self.t_rain_entry.pack(side='right')
		# -------- MODCOD Adaptativo (Fase 5) --------
		mod_frame = ttk.LabelFrame(left, text='Modulación / Coding (Fase 5)')
		mod_frame.pack(fill='x', pady=6)
		self.modcod_auto_var = tk.BooleanVar(value=self.core.params.get(["MODCOD","auto_default"]))
		self.modcod_selected_var = tk.StringVar(value=self.core.params.get(["MODCOD","default"]))
		mod_top = ttk.Frame(mod_frame); mod_top.pack(fill='x', padx=2, pady=1)
		mod_table = [e['name'] for e in self.core.params.get(["MODCOD","table"]) ]
		self.modcod_combo = ttk.Combobox(mod_top, values=mod_table, textvariable=self.modcod_selected_var, state='readonly', width=14)
		self.modcod_combo.pack(side='left')
		self.modcod_combo.bind('<<ComboboxSelected>>', lambda e: self.update_metrics())
		auto_chk = ttk.Checkbutton(mod_top, text='Auto', variable=self.modcod_auto_var, command=lambda: self.update_metrics())
		auto_chk.pack(side='right')
		# Parámetros dependientes de MODCOD (solo lectura)
		mod_mid1 = ttk.Frame(mod_frame); mod_mid1.pack(fill='x', padx=2, pady=1)
		self.rb_var = tk.DoubleVar(value=self.core.throughput['Rb_Mbps'])  # interno
		ttk.Label(mod_mid1, text='Rb (Mbps):', font=bold_lbl).pack(side='left')
		self.rb_value_label = ttk.Label(mod_mid1, text=f"{self.rb_var.get():.3f}", font=bold_lbl, width=8, anchor='e')
		self.rb_value_label.pack(side='right')
		mod_mid2 = ttk.Frame(mod_frame); mod_mid2.pack(fill='x', padx=2, pady=1)
		self.ebn0_req_var = tk.DoubleVar(value=self.core.throughput['EbN0_req_dB'])  # interno
		ttk.Label(mod_mid2, text='Eb/N0 Req (dB):', font=bold_lbl).pack(side='left')
		self.ebn0_req_value_label = ttk.Label(mod_mid2, text=f"{self.ebn0_req_var.get():.2f}", font=bold_lbl, width=8, anchor='e')
		self.ebn0_req_value_label.pack(side='right')
		mod_mid3 = ttk.Frame(mod_frame); mod_mid3.pack(fill='x', padx=2, pady=1)
		self.modcod_eff_label = ttk.Label(mod_mid3, text='Eff: -- b/Hz', font=bold_lbl)
		self.modcod_eff_label.pack(side='left')
		self.modcod_req_label = ttk.Label(mod_mid3, text='Eb/N0 Req: -- dB', font=bold_lbl)
		self.modcod_req_label.pack(side='right')
		mod_bot = ttk.Frame(mod_frame); mod_bot.pack(fill='x', padx=2, pady=1)
		self.modcod_status_label = ttk.Label(mod_bot, text='Estado: --', font=bold_lbl)
		self.modcod_status_label.pack(side='left')
		# -------- Latencias Detalladas (Fase 5) --------
		lat_frame = ttk.LabelFrame(left, text='Latencias (Fase 5)')
		lat_frame.pack(fill='x', pady=6)
		self.proc_lat_var = tk.DoubleVar(value=self.core.params.get(["Latencies","Processing_delay_ms"]))
		self.switch_lat_var = tk.DoubleVar(value=self.core.params.get(["Latencies","Switching_delay_ms"]))
		row_l1 = ttk.Frame(lat_frame); row_l1.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_l1, text='Proc (ms):', font=bold_lbl).pack(side='left')
		self.proc_entry = ttk.Entry(row_l1, textvariable=self.proc_lat_var, width=6); self.proc_entry.pack(side='right')
		row_l2 = ttk.Frame(lat_frame); row_l2.pack(fill='x', padx=2, pady=1)
		ttk.Label(row_l2, text='Switch (ms):', font=bold_lbl).pack(side='left')
		self.switch_entry = ttk.Entry(row_l2, textvariable=self.switch_lat_var, width=6); self.switch_entry.pack(side='right')
		# ---------------- Inputs Pérdidas (Fase 2) ----------------
		loss_frame = ttk.LabelFrame(left, text='Pérdidas (dB)')
		loss_frame.pack(fill='x', pady=6)
		self.loss_vars = {}
		loss_order = [
			('RFL_feeder','Feeder RF'),
			('AML_misalignment','Desalineación Antena'),
			('AA_atmos','Atenuación Atmosférica'),
			('Rain_att','Atenuación Lluvia'),
			('PL_polarization','Desajuste Polarización'),
			('L_pointing','Pérdida Apuntamiento'),
			('L_impl','Pérdidas Implementación'),
		]
		for key,label in loss_order:
			self.loss_vars[key] = tk.DoubleVar(value=self.core.losses[key])
			row_f = ttk.Frame(loss_frame); row_f.pack(fill='x', padx=2, pady=1)
			ttk.Label(row_f, text=label+':', font=bold_lbl).pack(side='left')
			ttk.Entry(row_f, textvariable=self.loss_vars[key], width=6).pack(side='right')
		self.run_btn = ttk.Button(left, text="Iniciar", command=self.toggle_run); self.run_btn.pack(anchor='w', pady=2)
		self.reset_btn = ttk.Button(left, text="Reset", command=self.reset); self.reset_btn.pack(anchor='w', pady=2)
		center = ttk.Frame(self.mainframe); center.pack(side='left', fill='both', expand=True)
		self.canvas = tk.Canvas(center, width=600, height=480, bg='white', highlightthickness=0); self.canvas.pack(fill='both', expand=True)
		slider_frame = ttk.Frame(center); slider_frame.pack(fill='x')
		
		# Control de modo de tiempo
		time_control_frame = ttk.Frame(slider_frame)
		time_control_frame.pack(fill='x', pady=2)
		self.manual_mode_var = tk.BooleanVar(value=False)
		manual_check = ttk.Checkbutton(time_control_frame, text='Modo Manual', variable=self.manual_mode_var, command=self._toggle_manual_mode)
		manual_check.pack(side='left')
		
		# Control de sensibilidad de tiempo
		sensitivity_frame = ttk.Frame(time_control_frame)
		sensitivity_frame.pack(side='right')
		ttk.Label(sensitivity_frame, text='Sensibilidad:', font=('Segoe UI', 9)).pack(side='left', padx=(10,2))
		self.time_sensitivity_var = tk.DoubleVar(value=1.0)  # Factor multiplicador de sensibilidad
		sensitivity_scale = tk.Scale(sensitivity_frame, from_=0.1, to=5.0, orient='horizontal', resolution=0.1,
									variable=self.time_sensitivity_var, length=120, showvalue=True)
		sensitivity_scale.pack(side='left')
		ttk.Label(time_control_frame, text='(Pausar para control fino de posiciones)', font=('Segoe UI', 9)).pack(side='left', padx=10)
		
		# Barra principal de tiempo de simulación
		self.time_slider_var = tk.DoubleVar(value=0.0)
		self.time_slider = tk.Scale(slider_frame, from_=0, to=self.max_simulation_time_s, orient='horizontal', resolution=0.1, 
									label='Tiempo Simulación [s] - Controla LEO + Tierra + GEO sincronizados', 
									variable=self.time_slider_var, command=lambda v: self._on_time_slider(float(v)))
		self.time_slider.pack(fill='x', pady=2)
		
		# Barras de ajuste fino (solo activas en modo manual)
		self.orbit_slider_var = tk.DoubleVar(value=self.orbit_angle_deg); self.user_adjusting_slider = False
		self.orbit_slider = tk.Scale(slider_frame, from_=0, to=359.9, orient='horizontal', resolution=0.1, 
									label='Ajuste Fino LEO [°] - Solo activo en Modo Manual', 
									variable=self.orbit_slider_var, command=lambda v: self._on_slider_change(float(v)))
		self.orbit_slider.pack(fill='x', pady=2)
		self.orbit_slider.configure(state='disabled')  # Inicialmente deshabilitado
		
		self.geo_slider_var = tk.DoubleVar(value=0.0)
		self.geo_slider = tk.Scale(slider_frame, from_=-180, to=180, orient='horizontal', resolution=0.5, 
								   label='Ajuste Fino GEO Long [°] - Solo activo en Modo Manual', 
								   variable=self.geo_slider_var, command=lambda v: self._on_geo_slider(float(v)))
		self.geo_slider.pack(fill='x', pady=2)
		self.geo_slider.configure(state='disabled')  # Inicialmente deshabilitado
		
		self.orbit_slider.bind('<ButtonPress-1>', lambda e: self._begin_slider()); self.orbit_slider.bind('<ButtonRelease-1>', lambda e: self._end_slider())
		self.canvas.bind('<Configure>', lambda e: self._draw_static())
		right = ttk.Frame(self.mainframe); right.pack(side='left', fill='y')
		# Panel de métricas desplazable (canvas + scrollbar)
		self.metrics_canvas = tk.Canvas(right, borderwidth=0, highlightthickness=0)
		self.metrics_scrollbar = ttk.Scrollbar(right, orient='vertical', command=self.metrics_canvas.yview)
		self.metrics_canvas.configure(yscrollcommand=self.metrics_scrollbar.set)
		self.metrics_canvas.pack(side='left', fill='both', expand=True)
		self.metrics_scrollbar.pack(side='right', fill='y')
		self.metrics_panel = ttk.Frame(self.metrics_canvas)
		self._metrics_window = self.metrics_canvas.create_window((0,0), window=self.metrics_panel, anchor='nw')
		self.metrics_panel.bind('<Configure>', lambda e: self.metrics_canvas.configure(scrollregion=self.metrics_canvas.bbox('all')))
		self.metrics_canvas.bind('<Configure>', lambda e: self.metrics_canvas.itemconfigure(self._metrics_window, width=e.width))
		# Soporte rueda ratón (Windows delta 120)
		self.metrics_canvas.bind_all('<MouseWheel>', self._on_mousewheel)
		self._init_metrics_table()
		# Botón colapsar pérdidas
		self.show_losses = False
		self.toggle_losses_btn = ttk.Button(right, text='Mostrar Pérdidas ▶', command=self._toggle_losses)
		self.toggle_losses_btn.pack(fill='x', pady=2)
		self.export_btn = ttk.Button(right, text='Exportar CSV/XLSX', command=self.export_csv); self.export_btn.pack(side='bottom', pady=4)

	def toggle_run(self):
		self.running = not self.running
		if self.running and self.start_time is None: 
			self.start_time = time.time()
			self.last_animation_time = None  # Reset para primera animación
		self.run_btn.config(text='Parar' if self.running else 'Iniciar')
		
		# Iniciar animación para ambos modos (LEO y GEO ahora ambos animan)
		if self.running: 
			self._animate()
		if not self.running: 
			self.orbit_slider_var.set(self.orbit_angle_deg)

	def reset(self):
		self.orbit_angle_deg = (360.0 - self.horizon_central_angle_deg)
		self.earth_rotation_angle_deg = 0.0
		self.geo_rotation_angle_deg = 0.0
		self.simulation_time_s = 0.0
		self.history.clear()
		self.start_time = None
		self.last_animation_time = None
		self.update_metrics(); self._draw_dynamic(); 
		self.orbit_slider_var.set(self.orbit_angle_deg); self.geo_slider_var.set(0.0); self.time_slider_var.set(0.0)

	def _begin_slider(self):
		if self.running:
			self.running = False; self.run_btn.config(text='Iniciar')
		self.user_adjusting_slider = True

	def _on_slider_change(self, val: float):
		if not self.user_adjusting_slider: return
		if not self.manual_time_control: return  # Solo en modo manual
		self.orbit_angle_deg = val % 360.0; self._draw_dynamic(); self.update_metrics()

	def _on_geo_slider(self, val: float):
		if not self.manual_time_control: return  # Solo en modo manual
		if self.mode_var.get() == 'GEO': self._draw_dynamic(); self.update_metrics()

	def _end_slider(self):
		self.user_adjusting_slider = False; self.orbit_slider_var.set(self.orbit_angle_deg)
	
	def _toggle_manual_mode(self):
		"""Cambia entre modo automático y manual."""
		self.manual_time_control = self.manual_mode_var.get()
		
		if self.manual_time_control:
			# Modo manual: habilitar barras de ajuste fino
			self.orbit_slider.configure(state='normal')
			if self.mode_var.get() == 'GEO':
				self.geo_slider.configure(state='normal')
			# Si está corriendo, pausarlo
			if self.running:
				self.toggle_run()
		else:
			# Modo automático: deshabilitar barras de ajuste fino
			self.orbit_slider.configure(state='disabled')
			self.geo_slider.configure(state='disabled')
	
	def _on_time_slider(self, time_s: float):
		"""Maneja el cambio en la barra de tiempo de simulación."""
		if not self.running:  # Solo permitir cambio manual cuando esté pausado
			self.simulation_time_s = time_s
			self._update_positions_from_time()
			self._draw_dynamic()
			self.update_metrics()
	
	def _update_positions_from_time(self):
		"""Actualiza todas las posiciones basándose en el tiempo de simulación."""
		# Aplicar factor de sensibilidad definido por usuario
		effective_time = self.simulation_time_s * self.time_sensitivity_var.get()
		
		# Calcular rotación de la Tierra basada en tiempo
		earth_rotation_rate_deg_per_s = EARTH_ROTATION_DEG_PER_S * self.time_scale_factor
		self.earth_rotation_angle_deg = (effective_time * earth_rotation_rate_deg_per_s) % 360.0
		
		# GEO siempre sincronizado con Tierra
		self.geo_rotation_angle_deg = self.earth_rotation_angle_deg
		
		# LEO: calcular posición orbital basada en tiempo
		if self.mode_var.get() == 'LEO':
			# Velocidad orbital real para LEO
			orbital_velocity_deg_per_s = math.sqrt(MU_EARTH / (self.orbit_r_km * 1000)) * 180 / (math.pi * self.orbit_r_km * 1000) * self.time_scale_factor
			orbit_change = effective_time * orbital_velocity_deg_per_s
			self.orbit_angle_deg = ((360.0 - self.horizon_central_angle_deg) + orbit_change) % 360.0

	def export_csv(self):
		"""Exporta la historia en CSV/XLSX estructurado por secciones de la interfaz.
		
		ESTRUCTURA ORGANIZADA POR SECCIONES DE LA GUI:
		- Parámetros Básicos, Enlaces Separados UL/DL, End-to-End, Potencia/Back-off,
		  Ruido/Rendimiento, Geometría Orbital, Doppler, Pérdidas
		- Para XLSX: cabeceras en NEGRITA, columnas ANCHAS, paneles congelados
		- Refleja exactamente la estructura de la interfaz
		"""
		if not self.history:
			self._append_metrics("No hay datos para exportar. Inicia la simulación primero.\n")
			return
		
		from tkinter import filedialog
		path = filedialog.asksaveasfilename(
			defaultextension='.csv', 
			filetypes=[('CSV','*.csv'), ('Excel','*.xlsx')], 
			title='Guardar resultados organizados por secciones'
		)
		if not path: return
		
		# ESTRUCTURA ORGANIZADA SEGÚN LA INTERFAZ GUI
		section_structure = {
			'parametros_basicos': {
				'title': '=== PARÁMETROS BÁSICOS ===',
				'fields': ['time_s', 'mode', 'elevation_deg', 'slant_range_km', 'fspl_db', 
						  'latency_ms_one_way', 'latency_total_rtt_ms', 'cn_quality']
			},
			'enlaces_uplink': {
				'title': '=== ENLACES SEPARADOS - UPLINK ===',
				'fields': ['ul_cn0_dbhz', 'ul_cn_db', 'ul_frequency_ghz', 'ul_bandwidth_mhz', 
						  'ul_gt_dbk', 'ul_cn_quality']
			},
			'enlaces_downlink': {
				'title': '=== ENLACES SEPARADOS - DOWNLINK ===',
				'fields': ['dl_cn0_dbhz', 'dl_cn_db', 'dl_frequency_ghz', 'dl_bandwidth_mhz', 
						  'dl_gt_dbk', 'dl_cn_quality']
			},
			'end_to_end': {
				'title': '=== END-TO-END ===',
				'fields': ['e2e_latency_total_ms_one_way', 'e2e_latency_total_rtt_ms', 'e2e_cn_total_db',
						  'e2e_cinr_total_db', 'e2e_worst_link', 'e2e_status']
			},
			'potencia_backoff': {
				'title': '=== POTENCIA Y BACK-OFF ===',
				'fields': ['eirp_sat_dbw', 'input_backoff_db', 'output_backoff_db', 'eirp_dbw']
			},
			'ruido_rendimiento': {
				'title': '=== RUIDO Y RENDIMIENTO ===',
				'fields': ['T_sys_K', 'N0_dBHz', 'EbN0_dB', 'EbN0_req_dB', 'Eb_margin_dB', 'modcod_status']
			},
			'geometria_orbital': {
				'title': '=== GEOMETRÍA ORBITAL ===',
				'fields': ['orbit_angle_deg', 'orbital_radius_km', 'orbital_velocity_kms', 
						  'angular_velocity_degs', 'range_rate_kms', 'orbital_period_min']
			},
			'doppler': {
				'title': '=== DOPPLER ===',
				'fields': ['doppler_instantaneous_khz', 'doppler_max_theoretical_khz']
			},
			'perdidas': {
				'title': '=== PÉRDIDAS ===',
				'fields': ['loss_total_extra_db', 'RFL_feeder', 'AML_misalignment', 'AA_atmos', 
						  'Rain_att', 'PL_polarization', 'L_pointing', 'L_impl']
			}
		}
		
		# MAPEO DE ETIQUETAS ORGANIZADO POR SECCIONES DE LA INTERFAZ
		enhanced_label_map = {
			# === PARÁMETROS BÁSICOS ===
			'time_s': 'TIEMPO [s]',
			'mode': 'MODO',
			'elevation_deg': 'ELEVACION [°]',
			'slant_range_km': 'DISTANCIA SLANT [km]',
			'fspl_db': 'FSPL (ESPACIO LIBRE) [dB]',
			'latency_ms_one_way': 'LATENCIA IDA [ms]',
			'latency_total_rtt_ms': 'LATENCIA RTT [ms]',
			'cn_quality': 'ESTADO C/N',
			
			# === ENLACES SEPARADOS - UPLINK ===
			'ul_cn0_dbhz': 'UL C/N0 [dBHz]',
			'ul_cn_db': 'UL C/N [dB]',
			'ul_frequency_ghz': 'UL FREQ [GHz]',
			'ul_bandwidth_mhz': 'UL BW [MHz]',
			'ul_gt_dbk': 'UL G/T [dB/K]',
			'ul_cn_quality': 'UL ESTADO C/N',
			
			# === ENLACES SEPARADOS - DOWNLINK ===
			'dl_cn0_dbhz': 'DL C/N0 [dBHz]',
			'dl_cn_db': 'DL C/N [dB]',
			'dl_frequency_ghz': 'DL FREQ [GHz]',
			'dl_bandwidth_mhz': 'DL BW [MHz]',
			'dl_gt_dbk': 'DL G/T [dB/K]',
			'dl_cn_quality': 'DL ESTADO C/N',
			
			# === END-TO-END ===
			'e2e_latency_total_ms_one_way': 'E2E LATENCIA TOTAL [ms]',
			'e2e_latency_total_rtt_ms': 'E2E LATENCIA RTT [ms]',
			'e2e_cn_total_db': 'E2E C/N TOTAL [dB]',
			'e2e_cinr_total_db': 'E2E CINR TOTAL [dB]',
			'e2e_worst_link': 'E2E ENLACE CRÍTICO',
			'e2e_status': 'E2E ESTADO',
			
			# === POTENCIA Y BACK-OFF ===
			'eirp_sat_dbw': 'EIRP SATURADO [dBW]',
			'input_backoff_db': 'BACK-OFF ENTRADA [dB]',
			'output_backoff_db': 'BACK-OFF SALIDA [dB]',
			'eirp_dbw': 'EIRP EFECTIVO [dBW]',
			
			# === RUIDO Y RENDIMIENTO ===
			'T_sys_K': 'TEMPERATURA SISTEMA T_SYS [K]',
			'N0_dBHz': 'DENSIDAD RUIDO N0 [dBHz]',
			'EbN0_dB': 'EB/N0 [dB]',
			'EbN0_req_dB': 'EB/N0 REQUERIDO [dB]',
			'Eb_margin_dB': 'MARGEN EB/N0 [dB]',
			'modcod_status': 'ESTADO MODCOD',
			
			# === GEOMETRÍA ORBITAL ===
			'orbit_angle_deg': 'ANGULO CENTRAL Δ [°]',
			'orbital_radius_km': 'RADIO ORBITAL [km]',
			'orbital_velocity_kms': 'VELOCIDAD ORBITAL [km/s]',
			'angular_velocity_degs': 'VELOCIDAD ANGULAR [°/s]',
			'range_rate_kms': 'RATE CAMBIO DISTANCIA [km/s]',
			'orbital_period_min': 'PERIODO ORBITAL [min]',
			
			# === DOPPLER ===
			'doppler_instantaneous_khz': 'DOPPLER INSTANTANEO [kHz]',
			'doppler_max_theoretical_khz': 'DOPPLER MÁX TEORICO [kHz]',
			
			# === PÉRDIDAS ===
			'loss_total_extra_db': 'Σ PERDIDAS EXTRA [dB]',
			'RFL_feeder': 'FEEDER RF [dB]',
			'AML_misalignment': 'DESALINEACION ANTENA [dB]',
			'AA_atmos': 'AA ATMOSFERICA [dB]',
			'Rain_att': 'ATENUACION LLUVIA [dB]',
			'PL_polarization': 'DESAJUSTE POLARIZACION [dB]',
			'L_pointing': 'PERDIDA APUNTAMIENTO [dB]',
			'L_impl': 'PERDIDAS IMPLEMENTACION [dB]'
		}
		
		# Construir el orden final de campos
		final_field_order = []
		for section_key, section_data in section_structure.items():
			final_field_order.extend(section_data['fields'])
		
		# Añadir campos adicionales no contemplados (compatibilidad)
		if self.history:
			for field in self.history[0].keys():
				if field not in final_field_order and field not in enhanced_label_map:
					final_field_order.append(field)
					enhanced_label_map[field] = field.upper()
		
		try:
			if path.lower().endswith('.xlsx'):
				try:
					from openpyxl import Workbook
					from openpyxl.styles import Font, PatternFill, Alignment
					from openpyxl.utils import get_column_letter
				except ImportError:
					self._append_metrics("openpyxl no instalado. Usa: pip install openpyxl\nExportando como CSV...\n")
					path = path.replace('.xlsx', '.csv')
				else:
					# CREACIÓN XLSX CON FORMATO PROFESIONAL
					wb = Workbook()
					ws = wb.active
					ws.title = 'Simulación LEO-GEO'
					
					# Escribir cabeceras organizadas por secciones
					headers = [enhanced_label_map.get(field, field.upper()) for field in final_field_order]
					ws.append(headers)
					
					# FORMATO PROFESIONAL DE CABECERAS - MEJORADO
					header_font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
					header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
					header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
					
					for col_num, cell in enumerate(ws[1], 1):
						cell.font = header_font
						cell.fill = header_fill
						cell.alignment = header_alignment
					
					# Escribir todos los datos de la simulación
					for row_data in self.history:
						ws.append([row_data.get(field, '') for field in final_field_order])
					
					# COLUMNAS ANCHAS Y BIEN FORMATEADAS
					for col_num in range(1, len(final_field_order) + 1):
						column_letter = get_column_letter(col_num)
						
						# Obtener el header para determinar ancho
						header_text = ws.cell(row=1, column=col_num).value or ""
						header_length = len(str(header_text))
						
						# COLUMNAS EXTRA ANCHAS para mejor legibilidad
						if header_length > 30:
							width = 40  # Columnas muy largas
						elif header_length > 25:
							width = 35  # Columnas largas
						elif header_length > 20:
							width = 30  # Columnas medianas
						elif header_length > 15:
							width = 25  # Columnas normales
						else:
							width = 20  # Columnas cortas, pero aún anchas
						
						ws.column_dimensions[column_letter].width = width
					
					# Altura de fila de cabecera MÁS ALTA
					ws.row_dimensions[1].height = 35
					
					# Congelar paneles para navegación fácil
					ws.freeze_panes = 'A2'
					
					# Guardar con formato completo
					wb.save(path)
					self._append_metrics(f"✅ XLSX exportado con formato profesional: {path}\n")
					self._append_metrics(f"📊 {len(final_field_order)} columnas organizadas por secciones\n")
					return
			
			# EXPORTACIÓN CSV ESTRUCTURADA (fallback)
			import csv
			with open(path, 'w', newline='', encoding='utf-8') as f:
				writer = csv.writer(f)
				
				# Escribir cabeceras organizadas
				headers = [enhanced_label_map.get(field, field.upper()) for field in final_field_order]
				writer.writerow(headers)
				
				# Escribir todos los datos capturados
				for row_data in self.history:
					writer.writerow([row_data.get(field, '') for field in final_field_order])
			
			self._append_metrics(f"✅ CSV exportado con estructura organizada: {path}\n")
			self._append_metrics(f"📊 {len(final_field_order)} columnas por secciones de la interfaz\n")
			
		except Exception as e:
			self._append_metrics(f"❌ Error exportando: {e}\n")

	def _append_metrics(self, text: str):
		# Por simplicidad mostramos mensajes emergentes en consola si no hay text widget
		print(text.rstrip())

	def _init_metrics_table(self):
		"""Crea etiquetas (nombre en negrita, valor coloreado)."""
		font_label = ('Segoe UI', 10, 'bold')
		font_value = ('Consolas', 11)
		# Definimos filas con secciones (None => separador visual)
		rows = [
			('— PARÁMETROS BÁSICOS —','section'),
			('Modo', '—'),
			('Tiempo Simulación [s]', '—'),
			('Modo Control', '—'),
			('Elevación [°]', '—'),
			('Distancia Slant [km]', '—'),
			('FSPL (Espacio Libre) [dB]', '—'),
			('Latencia Ida [ms]', '—'),
			('Latencia RTT [ms]', '—'),
			('C/N0 [dBHz]', '—'),
			('C/N [dB]', '—'),
			('Estado C/N', '—'),
			('G/T [dB/K]', '—'),
			('— POTENCIA Y BACK-OFF —','section'),
			('EIRP Saturado [dBW]', '—'),
			('Back-off Entrada [dB]', '—'),
			('Back-off Salida [dB]', '—'),
			('EIRP Efectivo [dBW]', '—'),
			('— RUIDO Y RENDIMIENTO —','section'),
			('Temperatura Sistema T_sys [K]', '—'),
			('Densidad Ruido N0 [dBHz]', '—'),
			('Eb/N0 [dB]', '—'),
			('Eb/N0 Requerido [dB]', '—'),
			('Margen Eb/N0 [dB]', '—'),
			('Capacidad Shannon [Mbps]', '—'),
			('Eficiencia Espectral Real [b/Hz]', '—'),
			('Utilización vs Shannon [%]', '—'),
			('— GEOMETRÍA ORBITAL —','section'),
			('Ángulo Central Δ [°]', '—'),
			('Radio Orbital [km]', '—'),
			('Velocidad Orbital [km/s]', '—'),
			('Velocidad Angular ω [°/s]', '—'),
			('Rate Cambio Distancia [km/s]', '—'),
			('Periodo Orbital [min]', '—'),
			('Tiempo Visibilidad Restante [s]', '—'),
			('— DOPPLER —','section'),
			('Doppler Instantáneo [kHz]', '—'),
			('Doppler Máx Teórico [kHz]', '—'),
			('— PÉRDIDAS —','section'),
			('Σ Pérdidas Extra [dB]', '—'),
			('Path Loss Total [dB]', '—'),
		]
		self.metric_labels = {}
		row_index = 0
		for name, val in rows:
			if val == 'section':
				sep = ttk.Separator(self.metrics_panel, orient='horizontal')
				sep.grid(row=row_index, column=0, columnspan=2, sticky='ew', pady=(6,2))
				lbl_section = ttk.Label(self.metrics_panel, text=name, font=('Segoe UI',9,'bold'), foreground='#555')
				lbl_section.grid(row=row_index+1, column=0, columnspan=2, sticky='w', padx=2)
				row_index += 2
				continue
			lbl = ttk.Label(self.metrics_panel, text=name+':', font=font_label, anchor='w')
			lbl.grid(row=row_index, column=0, sticky='w', padx=(2,4), pady=1)
			val_lbl = ttk.Label(self.metrics_panel, text=val, font=font_value, foreground='#004080', anchor='e')
			val_lbl.grid(row=row_index, column=1, sticky='e', padx=(4,6), pady=1)
			self.metric_labels[name] = val_lbl
			row_index += 1
		self.loss_rows_start_index = row_index
		# Filas detalladas de pérdidas (creadas pero ocultas inicialmente)
		loss_detail_rows = [
			('Feeder RF [dB]', 'RFL_feeder'),
			('Desalineación Antena [dB]', 'AML_misalignment'),
			('Atenuación Atmosférica [dB]', 'AA_atmos'),
			('Atenuación Lluvia [dB]', 'Rain_att'),
			('Desajuste Polarización [dB]', 'PL_polarization'),
			('Pérdida Apuntamiento [dB]', 'L_pointing'),
			('Pérdidas Implementación [dB]', 'L_impl'),
		]
		self.loss_detail_label_map = {}
		current_row = self.loss_rows_start_index
		for disp, key in loss_detail_rows:
			lbl = ttk.Label(self.metrics_panel, text=disp+':', font=font_label, anchor='w')
			val_lbl = ttk.Label(self.metrics_panel, text='—', font=font_value, foreground='#004080', anchor='e')
			# No grid todavía (colapsado inicialmente)
			self.loss_detail_label_map[key] = val_lbl
		self.metrics_panel.columnconfigure(0, weight=1)
		self.metrics_panel.columnconfigure(0, weight=1)
		self.metrics_panel.columnconfigure(1, weight=1)

	def _draw_static(self):
		"""Redibuja fondo recalculando escala para que GEO completo quepa.
		Solo dibuja elementos que NO rotan: órbitastraces, referencias fijas.
		
		Escala: se reserva ~90% del mínimo lado para el diámetro GEO.
		"""
		w = max(self.canvas.winfo_width(), 200); h = max(self.canvas.winfo_height(), 200); self.canvas.delete('all')
		min_side = min(w, h)
		# Margen: 5% alrededor => diámetro GEO ocupa 90% => radio GEO 45% del min_side
		max_geo_radius_px = 0.45 * min_side
		self.scale_px_per_km = max_geo_radius_px / self.geo_orbit_r_km
		self.earth_radius_px = self.Re_km * self.scale_px_per_km
		# Radio físico (escala) de LEO y GEO
		self.leo_orbit_r_px_physical = self.orbit_r_km * self.scale_px_per_km
		self.geo_orbit_r_px = self.geo_orbit_r_km * self.scale_px_per_km
		# Radio visual para LEO (solo estética). Lo separamos un porcentaje del gap Tierra-GEO.
		gap_total_px = self.geo_orbit_r_px - self.earth_radius_px
		visual_fraction = 0.18  # Ajustable: 0.0=pegado, 1.0=mitad del camino hacia GEO
		self.leo_orbit_r_px_visual = self.earth_radius_px + max(18, visual_fraction * gap_total_px)
		# Centro (ligero desplazamiento vertical para texto)
		self.cx = w / 2
		self.cy = h / 2 + h * 0.04
		
		# Dibujar referencias fijas (no rotativas)
		# Líneas de referencia orbital (opcionales)
		# self.canvas.create_oval(self.cx-self.leo_orbit_r_px_visual, self.cy-self.leo_orbit_r_px_visual, 
		#                        self.cx+self.leo_orbit_r_px_visual, self.cy+self.leo_orbit_r_px_visual, 
		#                        outline='#cccccc', dash=(3,4), tags='static')
		# self.canvas.create_oval(self.cx-self.geo_orbit_r_px, self.cy-self.geo_orbit_r_px, 
		#                        self.cx+self.geo_orbit_r_px, self.cy+self.geo_orbit_r_px, 
		#                        outline='#aaaaff', dash=(2,3), tags='static')
		
		self._draw_earth_and_surface()
		self._draw_dynamic()

	def _draw_earth_and_surface(self):
		"""Dibuja la Tierra y elementos de superficie que rotan con ella."""
		self.canvas.delete('earth')
		
		# Tierra (base circular)
		self.canvas.create_oval(self.cx-self.earth_radius_px, self.cy-self.earth_radius_px, 
		                       self.cx+self.earth_radius_px, self.cy+self.earth_radius_px, 
		                       fill='#e0f4ff', outline='#0077aa', width=2, tags='earth')
		
		# Marcas en la superficie terrestre que indican rotación
		self._draw_earth_surface_features()
		
		# Ground station y futuros elementos de superficie
		self._draw_surface_elements()

	def _draw_earth_surface_features(self):
		"""Dibuja características en la superficie terrestre para visualizar rotación."""
		# Dibujar líneas de longitud como referencia visual de rotación
		for i in range(8):  # 8 líneas cada 45 grados
			angle_deg = i * 45.0 + self.earth_rotation_angle_deg
			angle_rad = math.radians(angle_deg)
			# Línea desde centro hasta borde
			end_x = self.cx + self.earth_radius_px * 0.9 * math.sin(angle_rad)
			end_y = self.cy - self.earth_radius_px * 0.9 * math.cos(angle_rad)
			self.canvas.create_line(self.cx, self.cy, end_x, end_y, 
			                       fill='#4488cc', width=1, dash=(2,2), tags='earth')

	def _draw_surface_elements(self):
		"""Dibuja elementos en la superficie que rotan con la Tierra."""
		# Ground Station - posición fija relativa a la Tierra (rota con ella)
		gs_angle_rad = math.radians(self.earth_rotation_angle_deg)  # GS en longitud 0° inicialmente
		self.gs_x = self.cx + self.earth_radius_px * math.sin(gs_angle_rad)
		self.gs_y = self.cy - self.earth_radius_px * math.cos(gs_angle_rad)
		
		self.canvas.create_oval(self.gs_x-5, self.gs_y-5, self.gs_x+5, self.gs_y+5, 
		                       fill='green', outline='black', tags='earth')
		self.canvas.create_text(self.gs_x+10, self.gs_y-10, text='GS', anchor='w', tags='earth')
		
		# Dibujar jammers terrestres (rotan con la Tierra)
		self._draw_jammers()

	def _draw_jammers(self):
		"""Dibuja los jammers terrestres en el canvas."""
		if not hasattr(self, 'jammer_manager') or not self.jammer_manager:
			return
		
		# Obtener posiciones de jammers considerando rotación terrestre
		# Aproximación: GS en lat=0, lon=0 para simplificar
		gs_lat, gs_lon = 0.0, 0.0
		jammer_positions = self.jammer_manager.get_jammer_positions(
			gs_lat, gs_lon, self.earth_rotation_angle_deg
		)
		
		for jammer_pos in jammer_positions:
			# Calcular posición relativa al GS en píxeles
			# Escala aproximada: dx, dy están en km desde GS
			dx_km = jammer_pos['dx']
			dy_km = jammer_pos['dy']
			
			# Convertir a píxeles considerando la escala del canvas
			# Usar una escala reducida para que los jammers sean visibles
			jammer_scale = 0.1  # Los jammers se muestran más cerca para visibilidad
			dx_px = dx_km * self.scale_px_per_km * jammer_scale
			dy_px = dy_km * self.scale_px_per_km * jammer_scale
			
			# Posición final del jammer
			jammer_x = self.gs_x + dx_px
			jammer_y = self.gs_y - dy_px  # Invertir Y (canvas coordenadas)
			
			# Dibujar jammer como círculo rojo
			jammer_radius = 4
			self.canvas.create_oval(
				jammer_x - jammer_radius, jammer_y - jammer_radius,
				jammer_x + jammer_radius, jammer_y + jammer_radius,
				fill='red', outline='darkred', width=2, tags='earth'
			)
			
			# Etiqueta del jammer
			self.canvas.create_text(
				jammer_x + 8, jammer_y - 8, 
				text=jammer_pos['name'], anchor='w', 
				font=('Segoe UI', 8), fill='darkred', tags='earth'
			)
			
			# Línea de conexión desde GS (opcional, para mostrar relación)
			self.canvas.create_line(
				self.gs_x, self.gs_y, jammer_x, jammer_y,
				fill='red', width=1, dash=(2, 2), tags='earth'
			)

	def _change_mode(self):
		# Cargar parámetros específicos según modo
		self._load_link_presets_for_mode()
		
		if self.mode_var.get() == 'LEO':
			self.eirp_var.set(self.core.eirp_dbw); self.gt_var.set(self.core.gt_dbk)
			# Habilitar barras solo en modo manual
			if self.manual_time_control:
				self.orbit_slider.configure(state='normal')
			self.geo_slider.configure(state='disabled')
		else:
			self.eirp_var.set(self.core.geo_eirp_dbw); self.gt_var.set(self.core.geo_gt_dbk)
			self.orbit_slider.configure(state='disabled')
			# Habilitar barras solo en modo manual
			if self.manual_time_control:
				self.geo_slider.configure(state='normal')
		self.update_metrics(); self._draw_dynamic()

	def _draw_dynamic(self):
		"""Redibuja elementos dinámicos: satélites y sus enlaces. 
		También actualiza la rotación terrestre."""
		
		# Actualizar rotación terrestre en cada frame
		self._draw_earth_and_surface()
		
		# Limpiar elementos dinámicos previos
		self.canvas.delete('dyn')
		
		if self.mode_var.get() == 'LEO':
			# LEO: dibujar órbita visual y satélite
			orbit_r_px = self.leo_orbit_r_px_visual
			self.canvas.create_oval(self.cx-orbit_r_px, self.cy-orbit_r_px, self.cx+orbit_r_px, self.cy+orbit_r_px, outline='#cccccc', dash=(3,4), tags='dyn')
			phi = math.radians(self.orbit_angle_deg % 360.0)
			# Posición visual LEO
			sx = self.cx + orbit_r_px * math.sin(phi); sy = self.cy - orbit_r_px * math.cos(phi)
			
			# Cálculo geométrico CORREGIDO para LEO considerando rotación terrestre
			# Ángulo central = diferencia entre posición satélite y Ground Station
			satellite_angle_deg = self.orbit_angle_deg % 360.0
			gs_angle_deg = self.earth_rotation_angle_deg % 360.0
			
			# Calcular diferencia angular mínima (considerando que es un círculo)
			delta_raw = abs(satellite_angle_deg - gs_angle_deg)
			if delta_raw > 180:
				delta_raw = 360 - delta_raw
			delta_deg = delta_raw
			
			# Geometría física para elevación y distancia
			Re = self.Re_km; Ro = self.orbit_r_km
			delta_rad = math.radians(delta_deg)
			slant_km = math.sqrt(Re*Re + Ro*Ro - 2*Re*Ro*math.cos(delta_rad))
			if slant_km == 0:
				elev_deg = 90.0
			else:
				sin_e = (Ro * math.cos(delta_rad) - Re) / slant_km
				sin_e = max(-1.0, min(1.0, sin_e))
				elev_deg = math.degrees(math.asin(sin_e))
			visible = elev_deg > 0
			
			# Dibujar enlace LEO-GS
			self.canvas.create_line(self.gs_x, self.gs_y, sx, sy, fill=('red' if visible else '#bbbbbb'), dash=(5,4) if visible else (2,4), width=2 if visible else 1, tags='dyn')
			self.canvas.create_oval(sx-7, sy-7, sx+7, sy+7, fill='orange', outline='black', tags='dyn')
			self.canvas.create_text(sx+10, sy, text=f"LEO {elev_deg:.0f}°" + ("" if visible else " (OCULTO)"), anchor='w', tags='dyn')
			
			self.current_elevation_deg = elev_deg; self.current_visible = visible; self.current_slant_distance_m = slant_km * 1000.0
			self.current_delta_deg = delta_deg
		else:
			# GEO: dibujar órbita y satélite geoestacionario
			geo_r_px = self.geo_orbit_r_px
			self.canvas.create_oval(self.cx-geo_r_px, self.cy-geo_r_px, self.cx+geo_r_px, self.cy+geo_r_px, outline='#aaaaff', dash=(2,3), tags='dyn')
			
			# GEO posición: longitud relativa al GS + rotación síncrona con Tierra
			geo_longitude_relative = self.geo_slider_var.get()  # Longitud relativa del slider
			# Posición absoluta GEO = rotación tierra + longitud relativa
			geo_absolute_angle = self.earth_rotation_angle_deg + geo_longitude_relative
			phi_geo = math.radians(geo_absolute_angle)
			sx = self.cx + geo_r_px * math.sin(phi_geo); sy = self.cy - geo_r_px * math.cos(phi_geo)
			
			# Cálculo geométrico para GEO (ángulo entre GS y satélite GEO)
			gs_angle = math.radians(self.earth_rotation_angle_deg)
			geo_angle = math.radians(geo_absolute_angle)
			delta_long = abs(geo_angle - gs_angle)  # Diferencia angular
			
			Re = self.Re_km; Ro = self.geo_orbit_r_km
			slant_km = math.sqrt(Re*Re + Ro*Ro - 2*Re*Ro*math.cos(delta_long))
			if slant_km == 0:
				elev_deg = 90.0
			else:
				sin_e = (Ro * math.cos(delta_long) - Re) / slant_km
				sin_e = max(-1.0, min(1.0, sin_e))
				elev_deg = math.degrees(math.asin(sin_e))
			visible = elev_deg > 0
			
			# Dibujar enlace GEO-GS
			self.canvas.create_line(self.gs_x, self.gs_y, sx, sy, fill=('purple' if visible else '#bbbbbb'), dash=(5,3) if visible else (2,3), width=2 if visible else 1, tags='dyn')
			self.canvas.create_oval(sx-8, sy-8, sx+8, sy+8, fill='#6040ff', outline='black', tags='dyn')
			self.canvas.create_text(sx+10, sy, text=f"GEO {elev_deg:.0f}°", anchor='w', tags='dyn')
			
			self.current_elevation_deg = elev_deg; self.current_visible = visible; self.current_slant_distance_m = slant_km * 1000.0

	def _compute_slant_range_m(self, central_angle_deg: float) -> float:
		delta = math.radians(central_angle_deg); re = self.Re_km; ro = self.orbit_r_km; slant_km = math.sqrt(re*re + ro*ro - 2*re*ro*math.cos(delta)); return slant_km * 1000.0

	def _animate(self):
		"""Animación principal que actualiza posiciones de satélites y rotación terrestre."""
		if not self.running: 
			return
			
		# Calcular tiempo transcurrido desde última animación
		current_time = time.time()
		if self.last_animation_time is None:
			self.last_animation_time = current_time
			dt_s = 0.3  # Usar tiempo fijo para primer frame
		else:
			dt_s = current_time - self.last_animation_time
			self.last_animation_time = current_time
		
		# Aplicar factor de escalado temporal para visualización
		scaled_dt_s = dt_s * self.time_scale_factor
		
		# Actualizar tiempo de simulación
		self.simulation_time_s += scaled_dt_s
		if self.simulation_time_s > self.max_simulation_time_s:
			self.simulation_time_s = self.max_simulation_time_s
			# Opcional: pausar automáticamente al llegar al máximo
			# self.toggle_run()
		
		# Actualizar barra de tiempo
		self.time_slider_var.set(self.simulation_time_s)
		
		# Actualizar rotación terrestre (realista)
		earth_rotation_increment = EARTH_ROTATION_DEG_PER_S * scaled_dt_s
		self.earth_rotation_angle_deg = (self.earth_rotation_angle_deg + earth_rotation_increment) % 360.0
		
		# GEO rota sincrónicamente con la Tierra (geoestacionario)
		self.geo_rotation_angle_deg = self.earth_rotation_angle_deg
		
		if self.mode_var.get() == 'LEO':
			# LEO: calcular velocidad angular real y actualizar posición
			Re_m = EARTH_RADIUS_M
			ro_m = Re_m + self.core.constellation.satellites[0].altitude_m
			v_orbital_ms = math.sqrt(MU_EARTH / ro_m)  # Velocidad orbital en m/s
			omega_rad_s = v_orbital_ms / ro_m  # Velocidad angular en rad/s
			omega_deg_s = math.degrees(omega_rad_s)  # Velocidad angular en deg/s
			
			# Incremento LEO basado en dinámica orbital real
			leo_increment = omega_deg_s * scaled_dt_s
			self.orbit_angle_deg = (self.orbit_angle_deg + leo_increment) % 360.0
			
			# Actualizar slider si no está siendo manipulado por el usuario
			if not self.user_adjusting_slider: 
				self.orbit_slider_var.set(self.orbit_angle_deg)
		
		# Actualizar métricas y dibujo
		self.update_metrics()
		self._draw_dynamic()
		
		# Programar siguiente frame con intervalo configurado
		self.root.after(self.animation_interval_ms, self._animate)


	# ----------------------------- BLOQUES MODULARES (Fases 0-1) ----------------------------- #
	def update_metrics(self):
		"""Orquesta el refresco: parámetros -> geometría -> doppler -> enlace -> render tabla."""
		self._update_core_params()
		self._update_geometry_block()  # Actualizar geometría primero
		self._update_link_params()     # Después actualizar enlaces (usa geometría actualizada)
		self._update_doppler_block()
		self._update_link_block()
		self._update_latency_block()  # Fase 5
		# Primero seleccionamos MODCOD (deriva Rb y Eb/N0 requerido) y luego calculamos performance real
		self._update_modcod_block()  # Fase 5 (genera Rb_Mbps & EbN0_req)
		self._update_performance_block()  # Fase 4 (usa Rb derivado)
		self._render_metrics()
		self._append_history_row()
		self._update_link_gui()  # Actualiza GUI de los enlaces UL/DL

	def _update_core_params(self):
		self.core.eirp_dbw = float(self.eirp_var.get())
		self.core.gt_dbk = float(self.gt_var.get())
		# Usar frecuencia del tab activo UL/DL/E2E
		if hasattr(self, 'current_link_sense'):
			if self.current_link_sense == 'UL' and hasattr(self, 'ul_freq_var'):
				self.core.calc.frequency_hz = float(self.ul_freq_var.get()) * 1e9
			elif self.current_link_sense == 'DL' and hasattr(self, 'dl_freq_var'):
				self.core.calc.frequency_hz = float(self.dl_freq_var.get()) * 1e9
			elif self.current_link_sense == 'E2E':
				# Para End-to-End, usar frecuencia de referencia
				ref_link = self.bw_ref_var.get() if hasattr(self, 'bw_ref_var') else 'UL'
				if ref_link == 'DL' and hasattr(self, 'dl_freq_var'):
					self.core.calc.frequency_hz = float(self.dl_freq_var.get()) * 1e9
				else:  # UL por defecto
					self.core.calc.frequency_hz = float(self.ul_freq_var.get()) * 1e9 if hasattr(self, 'ul_freq_var') else 14.0e9
			else:
				# Usar UL por defecto
				self.core.calc.frequency_hz = 14.0e9  # UL frequency por defecto
		else:
			self.core.calc.frequency_hz = 14.0e9  # UL frequency por defecto
		self.core.calc.default_bandwidth_hz = float(self.bw_var.get()) * 1e6
		# Sincroniza GEO calc frecuencia/BW
		self.core.geo_calc.frequency_hz = self.core.calc.frequency_hz
		self.core.geo_calc.default_bandwidth_hz = self.core.calc.default_bandwidth_hz
		self._update_power_block()

	def _update_power_block(self):
		"""Fase 3: backoff y EIRP efectivo."""
		try:
			eirp_sat = float(self.eirp_sat_var.get())
		except Exception:
			eirp_sat = self.core.power.get('EIRP_sat_saturated', self.core.eirp_dbw)
		try:
			input_bo = max(0.0, float(self.input_bo_var.get()))
		except Exception:
			input_bo = 0.0
		output_bo = input_bo - 5.0 if input_bo > 0 else 0.0
		if self.manual_override_var.get():
			try:
				self.core.eirp_dbw = float(self.manual_eirp_var.get())
			except Exception:
				pass
		else:
			self.core.eirp_dbw = eirp_sat - input_bo
		self.core.power['EIRP_sat_saturated'] = eirp_sat
		self.core.power['Input_backoff'] = input_bo
		self.power_metrics = {
			'eirp_sat': eirp_sat,
			'input_bo': input_bo,
			'output_bo': output_bo,
			'eirp_eff': self.core.eirp_dbw,
			'manual_override': int(self.manual_override_var.get())
		}
		self.output_bo_label.config(text=f"Back-off Salida: {output_bo:.1f} dB")
		self.eirp_eff_label.config(text=f"EIRP Efectivo: {self.core.eirp_dbw:.1f} dBW")

	def _update_geometry_block(self):
		"""Calcula métricas geométricas y dinámica orbital ideal (solo LEO)."""
		if not hasattr(self, 'current_slant_distance_m'):
			self._draw_dynamic()
		mode = self.mode_var.get()
		self.geom: Dict[str, Any] = {k: float('nan') for k in ['delta_deg','r_orb_km','v_orb_kms','omega_deg_s','range_rate_kms','t_orb_min','visibility_remaining_s']}
		if mode == 'LEO':
			# Orbital (ideal circular)
			alt_m = self.core.constellation.satellites[0].altitude_m
			r_orb_m = EARTH_RADIUS_M + alt_m
			v_orb = math.sqrt(MU_EARTH / r_orb_m)  # m/s
			omega_rad_s = v_orb / r_orb_m
			omega_deg_s = math.degrees(omega_rad_s)
			T_orb_s = 2 * math.pi * math.sqrt(r_orb_m ** 3 / MU_EARTH)
			T_orb_min = T_orb_s / 60.0
			delta_deg = getattr(self, 'current_delta_deg', float('nan'))
			# Range rate analítica con signo según variación delta
			if not hasattr(self, '_prev_delta_deg'):
				self._prev_delta_deg = delta_deg
			approaching = delta_deg < self._prev_delta_deg  # se acerca a nadir
			delta_rad = math.radians(delta_deg)
			Re = EARTH_RADIUS_M
			r_orb = r_orb_m
			d_slant = self.current_slant_distance_m
			if d_slant <= 0:
				range_rate_ms = 0.0
			else:
				dd_dDelta = (Re * r_orb * math.sin(delta_rad)) / d_slant
				range_rate_ms = dd_dDelta * omega_rad_s
			if approaching:
				range_rate_ms *= -1.0
			self._prev_delta_deg = delta_deg
			# Tiempo de visibilidad restante (hasta horizonte)
			if self.current_visible:
				rem_deg = max(0.0, self.horizon_central_angle_deg - delta_deg)
				visibility_remaining_s = rem_deg / omega_deg_s if omega_deg_s > 0 else float('nan')
			else:
				visibility_remaining_s = 0.0
			self.geom.update({
				'delta_deg': delta_deg,
				'r_orb_km': r_orb_m / 1000.0,
				'v_orb_kms': v_orb / 1000.0,
				'omega_deg_s': omega_deg_s,
				'range_rate_kms': range_rate_ms / 1000.0,
				't_orb_min': T_orb_min,
				'visibility_remaining_s': visibility_remaining_s,
			})

	def _update_doppler_block(self):
		"""Doppler instantáneo y máximo (solo LEO)."""
		self.doppler: Dict[str, Any] = {'fd_hz': float('nan'), 'fd_max_hz': float('nan')}
		if self.mode_var.get() == 'LEO' and self.current_visible:
			f_c = self.core.calc.frequency_hz
			v_rad_ms = self.geom.get('range_rate_kms', float('nan')) * 1000.0
			v_orb_ms = self.geom.get('v_orb_kms', float('nan')) * 1000.0
			if not math.isnan(v_rad_ms):
				fd = (v_rad_ms / SPEED_OF_LIGHT) * f_c
				fd_max = (v_orb_ms / SPEED_OF_LIGHT) * f_c
				self.doppler.update({'fd_hz': fd, 'fd_max_hz': fd_max})

	def _update_link_block(self):
		"""FSPL, latencia y C/N actuales (bloque ya existente)."""
		mode = self.mode_var.get()
		calc = self.core.calc if mode == 'LEO' else self.core.geo_calc
		d = getattr(self, 'current_slant_distance_m', float('nan'))
		visible = getattr(self, 'current_visible', False)
		# Actualiza pérdidas desde inputs UI
		if hasattr(self, 'loss_vars'):
			for k,var in self.loss_vars.items():
				try:
					self.core.losses[k] = float(var.get())
				except Exception:
					self.core.losses[k] = 0.0
		loss_total_extra = sum(self.core.losses.values())
		if visible:
			fspl = calc.free_space_path_loss_db(d)
			lat_ow = calc.propagation_delay_ms(d)
			# Usar Path Loss Total (= FSPL + pérdidas extra) para degradar C/N0
			path_loss_total = fspl + loss_total_extra
			cn0 = self.core.eirp_dbw + self.core.gt_dbk - path_loss_total + calc.K_BOLTZ_TERM
			cn = calc.cn_db(cn0)
		else:
			fspl = float('nan'); lat_ow = float('nan'); cn0 = float('nan'); cn = float('nan'); path_loss_total = float('nan')
		self.link_metrics = {
			'fspl_db': fspl,
			'latency_ms_ow': lat_ow,
			'cn0_dbhz': cn0,
			'cn_db': cn,
			'loss_total_extra_db': loss_total_extra,
			'path_loss_total_db': path_loss_total,
		}

	def _update_latency_block(self):
		"""Fase 5: latencias totales (propagación + procesamiento + switching)."""
		try:
			proc_ms = max(0.0, float(self.proc_lat_var.get()))
		except Exception:
			proc_ms = 0.0
		try:
			switch_ms = max(0.0, float(self.switch_lat_var.get()))
		except Exception:
			switch_ms = 0.0
		self.core.latencies['Processing_delay_ms'] = proc_ms
		self.core.latencies['Switching_delay_ms'] = switch_ms
		prop_ow = self.link_metrics.get('latency_ms_ow', float('nan'))
		if not math.isnan(prop_ow):
			total_ow = prop_ow + proc_ms + switch_ms
			total_rtt = 2*prop_ow + 2*(proc_ms + switch_ms)
			self.link_metrics['total_latency_ow_ms'] = total_ow
			self.link_metrics['total_latency_rtt_ms'] = total_rtt

	def _update_modcod_block(self):
		"""Fase 5: selección adaptativa de MODCOD y actualización de Rb/EbN0_req."""
		mod_table = self.core.params.get(["MODCOD","table"])
		# Construir mapa nombre->entry con eficiencia
		best = None
		current_ebn0 = None
		# Usar Eb/N0 del tab activo en lugar del sistema tradicional
		current_ebn0 = self._get_active_ebn0_db()
		# Calcular eficiencias
		for entry in mod_table:
			bps = entry['bits_per_symbol'] * entry['code_rate']  # bits útiles por símbolo
			entry['efficiency_bps_hz'] = bps  # asumimos Nyquist 1 símbolo/Hz
		# Auto selección
		if self.modcod_auto_var.get() and current_ebn0 is not None and not math.isnan(current_ebn0):
			hyst = self.core.params.get(["MODCOD","hysteresis_db"])
			candidates = []
			for e in mod_table:
				if current_ebn0 - e['ebn0_req_db'] >= hyst:
					candidates.append(e)
			if candidates:
				best = max(candidates, key=lambda x: x['efficiency_bps_hz'])
			else:
				# Ninguna cumple margen; escoger la más robusta (menor req)
				best = min(mod_table, key=lambda x: x['ebn0_req_db'])
			self.modcod_selected_var.set(best['name'])
		else:
			# Manual: encontrar entry activo
			for e in mod_table:
				if e['name'] == self.modcod_selected_var.get():
					best = e; break
		if not best:
			best = mod_table[0]
		self.modcod_active = best
		# Actualizar requerimiento Eb/N0 y Rb
		self.core.throughput['EbN0_req_dB'] = best['ebn0_req_db']
		# Recalcular Rb en función de BW y eficiencia modcod (bits/Hz)
		bw_hz = self.core.calc.default_bandwidth_hz
		Rb_bps = best['efficiency_bps_hz'] * bw_hz
		self.core.throughput['Rb_Mbps'] = Rb_bps / 1e6
		# Actualizar variables y labels (siempre solo lectura)
		self.rb_var.set(self.core.throughput['Rb_Mbps'])
		self.ebn0_req_var.set(self.core.throughput['EbN0_req_dB'])
		if hasattr(self,'rb_value_label'):
			self.rb_value_label.config(text=f"{self.core.throughput['Rb_Mbps']:.3f}")
		if hasattr(self,'ebn0_req_value_label'):
			self.ebn0_req_value_label.config(text=f"{self.core.throughput['EbN0_req_dB']:.2f}")
		# Margen MODCOD
		if current_ebn0 is not None and not math.isnan(current_ebn0):
			self.modcod_margin_db = current_ebn0 - best['ebn0_req_db']
			if self.modcod_margin_db > 3:
				self.modcod_status = 'Excelente'
			elif self.modcod_margin_db >= 1:
				self.modcod_status = 'Aceptable'
			elif self.modcod_margin_db >= 0:
				self.modcod_status = 'Crítico'
			else:
				self.modcod_status = 'Insuficiente'
		else:
			self.modcod_margin_db = float('nan')
			self.modcod_status = '—'
		# Actualizar labels MODCOD
		self.modcod_eff_label.config(text=f"Eff: {best['efficiency_bps_hz']:.3f} b/Hz")
		self.modcod_req_label.config(text=f"Eb/N0 Req: {best['ebn0_req_db']:.2f} dB")
		color_map = {'Excelente':'#007700','Aceptable':'#c08000','Crítico':'#b05000','Insuficiente':'#b00000','—':'#666666'}
		self.modcod_status_label.config(text=f"Estado: {self.modcod_status}", foreground=color_map.get(self.modcod_status,'#444444'))

	def _add_dynamic_metric_row(self, display_name: str):
		"""Inserta una nueva fila de métrica al final si no existía."""
		if display_name in self.metric_labels:
			return
		# Buscar fila máxima actual
		max_row = 0
		for child in self.metrics_panel.grid_slaves():
			try:
				max_row = max(max_row, int(child.grid_info().get('row',0)))
			except Exception:
				pass
		row = max_row + 1
		font_label = ('Segoe UI', 10, 'bold'); font_value = ('Consolas', 11)
		lbl = ttk.Label(self.metrics_panel, text=display_name+':', font=font_label, anchor='w')
		lbl.grid(row=row, column=0, sticky='w', padx=(2,4), pady=1)
		val_lbl = ttk.Label(self.metrics_panel, text='—', font=font_value, foreground='#004080', anchor='e')
		val_lbl.grid(row=row, column=1, sticky='e', padx=(4,6), pady=1)
		self.metric_labels[display_name] = val_lbl

	def _update_performance_block(self):
		"""Fase 4: calcula T_sys, N0, Eb/N0, margen y métricas de capacidad.

		T_sys = suma de temperaturas componentes (simplificado).
		N0_dBHz = 10log10(k*T_sys) = -198.6 + 10log10(T_sys) (porque 228.6 = 10log10(1/k)).
		Eb/N0 = C/N0 - 10log10(Rb) (Rb en bit/s)
		Capacidad Shannon usando C/N lineal (=10^(C/N/10)).
		"""
		# Leer inputs ruido / throughput
		try:
			self.core.noise['T_rx'] = max(0.0, float(self.t_rx_var.get()))
		except Exception:
			self.core.noise['T_rx'] = 0.0
		try:
			self.core.noise['T_clear_sky'] = max(0.0, float(self.t_sky_var.get()))
		except Exception:
			self.core.noise['T_clear_sky'] = 0.0
		try:
			self.core.noise['T_rain_excess'] = max(0.0, float(self.t_rain_var.get()))
		except Exception:
			self.core.noise['T_rain_excess'] = 0.0
		T_sys = self.core.noise['T_rx'] + self.core.noise['T_clear_sky'] + self.core.noise['T_rain_excess']
		if T_sys <= 0: T_sys = float('nan')
		# N0: kT => 10log10(k) ~ -228.6 dBW/Hz; pero usamos C/N0 fórmula eirp+G/T - L + 228.6 => N0_dBHz = - (G/T - 10log10(T_sys))? Para simplicidad educativa usamos:
		# N0_dBHz (dBW/Hz) = -228.6 + 10log10(T_sys)  (valor absoluto de densidad de ruido)
		if not math.isnan(T_sys):
			N0_dBHz = -228.6 + 10*math.log10(T_sys)
		else:
			N0_dBHz = float('nan')
		# Throughput
		# (Rb y Eb/N0 Req ya fueron fijados por el bloque MODCOD; no se leen entradas del usuario)
		Rb_bps = self.core.throughput['Rb_Mbps'] * 1e6
		cn0 = self.link_metrics.get('cn0_dbhz', float('nan'))
		if not math.isnan(cn0) and Rb_bps > 0:
			EbN0_dB = cn0 - 10*math.log10(Rb_bps)
		else:
			EbN0_dB = float('nan')
		EbN0_req = self.core.throughput['EbN0_req_dB']
		margin_dB = EbN0_dB - EbN0_req if not math.isnan(EbN0_dB) else float('nan')
		# Capacidad Shannon
		cn_db = self.link_metrics.get('cn_db', float('nan'))
		if not math.isnan(cn_db):
			cn_lin = 10**(cn_db/10.0)
			BW_hz = self.core.calc.default_bandwidth_hz
			C_shannon_bps = BW_hz * math.log2(1+cn_lin)
			C_shannon_Mbps = C_shannon_bps / 1e6
			eta_shannon = (C_shannon_bps / BW_hz) if BW_hz>0 else float('nan')  # bits/s/Hz
			eta_real = (Rb_bps / BW_hz) if BW_hz>0 else float('nan')
			util_pct = (eta_real / eta_shannon * 100.0) if eta_shannon and not math.isnan(eta_shannon) and eta_shannon>0 else float('nan')
		else:
			C_shannon_Mbps = float('nan'); eta_shannon = float('nan'); eta_real = float('nan'); util_pct = float('nan')
		self.perf_metrics = {
			'T_sys_K': T_sys,
			'N0_dBHz': N0_dBHz,
			'EbN0_dB': EbN0_dB,
			'EbN0_req_dB': EbN0_req,
			'Eb_margin_dB': margin_dB,
			'Shannon_capacity_Mbps': C_shannon_Mbps,
			'Spectral_eff_real_bps_hz': eta_real,
			'Utilization_pct': util_pct,
		}

	def _assess_eb_margin(self, margin_db: float):
		if isinstance(margin_db, float) and math.isnan(margin_db):
			return ("—", '#666666')
		if margin_db > 3.0:
			return ("OK", '#007700')
		elif margin_db >= 0.0:
			return ("Justo", '#c08000')
		else:
			return ("Insuficiente", '#b00000')

	def _render_metrics(self):
		def fmt(v, pattern="{:.2f}"):
			return '—' if (isinstance(v, float) and math.isnan(v)) else (pattern.format(v) if isinstance(v, (int, float)) else str(v))
		visible = getattr(self, 'current_visible', False)
		# Básicos
		self.metric_labels['Modo'].config(text=self.mode_var.get())
		self.metric_labels['Tiempo Simulación [s]'].config(text=f"{self.simulation_time_s:.1f}")
		control_mode = "Manual" if self.manual_time_control else "Automático"
		self.metric_labels['Modo Control'].config(text=control_mode)
		elv_txt = f"{self.current_elevation_deg:.1f} ({'OK' if visible else 'OCULTO'})"
		self.metric_labels['Elevación [°]'].config(text=elv_txt, foreground=('#004080' if visible else '#aa0000'))
		self.metric_labels['Distancia Slant [km]'].config(text=fmt(self.current_slant_distance_m/1000.0, "{:.0f}"))
		self.metric_labels['FSPL (Espacio Libre) [dB]'].config(text=fmt(self._get_active_fspl_db()))
		self.metric_labels['Latencia Ida [ms]'].config(text=fmt(self.link_metrics['latency_ms_ow']))
		rtt_ms = self.link_metrics['latency_ms_ow'] * 2 if not math.isnan(self.link_metrics['latency_ms_ow']) else float('nan')
		self.metric_labels['Latencia RTT [ms]'].config(text=fmt(rtt_ms))
		# Fase 5: latencias totales
		if 'total_latency_ow_ms' in self.link_metrics:
			# Añadimos/actualizamos dinámicamente labels si no existen
			if 'Latencia Total Ida [ms]' not in self.metric_labels:
				self._add_dynamic_metric_row('Latencia Total Ida [ms]')
			if 'Latencia Total RTT [ms]' not in self.metric_labels:
				self._add_dynamic_metric_row('Latencia Total RTT [ms]')
			self.metric_labels['Latencia Total Ida [ms]'].config(text=fmt(self.link_metrics['total_latency_ow_ms']))
			self.metric_labels['Latencia Total RTT [ms]'].config(text=fmt(self.link_metrics['total_latency_rtt_ms']))
		else:
			# Si invisibles y existen, mostrar guion
			if 'Latencia Total Ida [ms]' in self.metric_labels:
				self.metric_labels['Latencia Total Ida [ms]'].config(text='—')
			if 'Latencia Total RTT [ms]' in self.metric_labels:
				self.metric_labels['Latencia Total RTT [ms]'].config(text='—')
		self.metric_labels['C/N0 [dBHz]'].config(text=fmt(self._get_active_cn0_dbhz()))
		self.metric_labels['C/N [dB]'].config(text=fmt(self._get_active_cn_db()))
		self.metric_labels['G/T [dB/K]'].config(text=f"{self.core.gt_dbk:.1f}")
		status_txt, color = self._assess_cn(self._get_active_cn_db())
		self.metric_labels['Estado C/N'].config(text=status_txt, foreground=color)

		# Potencia / Backoff
		self.metric_labels['EIRP Saturado [dBW]'].config(text=f"{self.power_metrics.get('eirp_sat', float('nan')):.1f}")
		self.metric_labels['Back-off Entrada [dB]'].config(text=f"{self.power_metrics.get('input_bo', float('nan')):.1f}")
		self.metric_labels['Back-off Salida [dB]'].config(text=f"{self.power_metrics.get('output_bo', float('nan')):.1f}")
		self.metric_labels['EIRP Efectivo [dBW]'].config(text=f"{self.core.eirp_dbw:.1f}")
		# Geometría
		self.metric_labels['Ángulo Central Δ [°]'].config(text=fmt(self.geom['delta_deg'], "{:.2f}"))
		self.metric_labels['Radio Orbital [km]'].config(text=fmt(self.geom['r_orb_km'], "{:.0f}"))
		self.metric_labels['Velocidad Orbital [km/s]'].config(text=fmt(self.geom['v_orb_kms'], "{:.2f}"))
		self.metric_labels['Velocidad Angular ω [°/s]'].config(text=fmt(self.geom['omega_deg_s'], "{:.3f}"))
		self.metric_labels['Rate Cambio Distancia [km/s]'].config(text=fmt(self.geom['range_rate_kms'], "{:.3f}"))
		self.metric_labels['Periodo Orbital [min]'].config(text=fmt(self.geom['t_orb_min'], "{:.1f}"))
		self.metric_labels['Tiempo Visibilidad Restante [s]'].config(text=fmt(self.geom['visibility_remaining_s'], "{:.1f}"))
		# Doppler
		fd_khz = self.doppler['fd_hz'] / 1e3 if not math.isnan(self.doppler['fd_hz']) else float('nan')
		fdmax_khz = self.doppler['fd_max_hz'] / 1e3 if not math.isnan(self.doppler['fd_max_hz']) else float('nan')
		self.metric_labels['Doppler Instantáneo [kHz]'].config(text=fmt(fd_khz, "{:.1f}"))
		self.metric_labels['Doppler Máx Teórico [kHz]'].config(text=fmt(fdmax_khz, "{:.1f}"))
		# Pérdidas
		self.metric_labels['Σ Pérdidas Extra [dB]'].config(text=fmt(self.link_metrics.get('loss_total_extra_db', float('nan'))))
		self.metric_labels['Path Loss Total [dB]'].config(text=fmt(self.link_metrics.get('path_loss_total_db', float('nan'))))
		# Ruido y rendimiento
		if hasattr(self, 'perf_metrics'):
			self.metric_labels['Temperatura Sistema T_sys [K]'].config(text=fmt(self.perf_metrics['T_sys_K'], "{:.1f}"))
			self.metric_labels['Densidad Ruido N0 [dBHz]'].config(text=fmt(self.perf_metrics['N0_dBHz'], "{:.1f}"))
			self.metric_labels['Eb/N0 [dB]'].config(text=fmt(self.perf_metrics['EbN0_dB'], "{:.2f}"))
			self.metric_labels['Eb/N0 Requerido [dB]'].config(text=f"{self.perf_metrics['EbN0_req_dB']:.2f}")
			self.metric_labels['Margen Eb/N0 [dB]'].config(text=fmt(self.perf_metrics['Eb_margin_dB'], "{:.2f}"), foreground=self._assess_eb_margin(self.perf_metrics['Eb_margin_dB'])[1])
			self.metric_labels['Capacidad Shannon [Mbps]'].config(text=fmt(self.perf_metrics['Shannon_capacity_Mbps'], "{:.2f}"))
			self.metric_labels['Eficiencia Espectral Real [b/Hz]'].config(text=fmt(self.perf_metrics['Spectral_eff_real_bps_hz'], "{:.3f}"))
			self.metric_labels['Utilización vs Shannon [%]'].config(text=fmt(self.perf_metrics['Utilization_pct'], "{:.1f}"))
		if getattr(self, 'show_losses', False) and hasattr(self, 'loss_detail_label_map'):
			for k,v in self.core.losses.items():
				if k in self.loss_detail_label_map:
					self.loss_detail_label_map[k].config(text=fmt(v, "{:.2f}"))

	def _assess_cn(self, cn_db: float):
		if isinstance(cn_db, float) and math.isnan(cn_db):
			return ("No visible", "#666666")
		try:
			val = float(cn_db)
		except Exception:
			return ("—", "#666666")
		if val > 15.0:
			return ("Excelente", "#007700")
		elif val >= 6.0:
			return ("Aceptable", "#c08000")
		else:
			return ("Crítico", "#b00000")

	def _append_history_row(self):
		"""Captura completa de datos organizados por secciones de la interfaz."""
		if self.running and self.start_time is not None:
			elapsed = time.time() - self.start_time
			visible = getattr(self, 'current_visible', False)
			fspl = self.link_metrics['fspl_db']
			lat_ow = self.link_metrics['latency_ms_ow']
			cn0 = self.link_metrics['cn0_dbhz']
			cn = self.link_metrics['cn_db']
			
			# === PARÁMETROS BÁSICOS ===
			row = {
				'time_s': round(elapsed, 3),
				'mode': self.mode_var.get(),
				'elevation_deg': round(self.current_elevation_deg, 2),
				'slant_range_km': round(self.current_slant_distance_m / 1000.0, 2),
				'fspl_db': None if math.isnan(fspl) else round(fspl, 2),
				'latency_ms_one_way': None if math.isnan(lat_ow) else round(lat_ow, 3),
				'latency_total_rtt_ms': round(self.link_metrics.get('total_latency_rtt_ms', lat_ow * 2), 3) if lat_ow and not math.isnan(lat_ow) else None,
				'cn_quality': self._assess_cn(cn)[0] if cn is not None else 'No visible',
			}
			
			# === ENLACES SEPARADOS - UPLINK & DOWNLINK ===
			if hasattr(self, 'link_out') and hasattr(self, 'link_state'):
				ul_out = self.link_out.get('UL')
				dl_out = self.link_out.get('DL')
				ul_state = self.link_state.get('UL')
				dl_state = self.link_state.get('DL')
				
				# UPLINK
				if ul_out and ul_state:
					ul_cn_quality = self._assess_cn(ul_out.CN_dB)[0] if ul_out.CN_dB is not None else 'No visible'
					row.update({
						'ul_cn0_dbhz': round(ul_out.CN0_dBHz, 2),
						'ul_cn_db': round(ul_out.CN_dB, 2),
						'ul_frequency_ghz': round(ul_state.f_Hz / 1e9, 4),
						'ul_bandwidth_mhz': round(ul_state.B_Hz / 1e6, 3),
						'ul_gt_dbk': round(ul_state.GT_dBK, 2),
						'ul_cn_quality': ul_cn_quality,
					})
				
				# DOWNLINK
				if dl_out and dl_state:
					dl_cn_quality = self._assess_cn(dl_out.CN_dB)[0] if dl_out.CN_dB is not None else 'No visible'
					row.update({
						'dl_cn0_dbhz': round(dl_out.CN0_dBHz, 2),
						'dl_cn_db': round(dl_out.CN_dB, 2),
						'dl_frequency_ghz': round(dl_state.f_Hz / 1e9, 4),
						'dl_bandwidth_mhz': round(dl_state.B_Hz / 1e6, 3),
						'dl_gt_dbk': round(dl_state.GT_dBK, 2),
						'dl_cn_quality': dl_cn_quality,
					})
				
				# === END-TO-END ===
				if ul_out and dl_out:
					combined = combine_end_to_end(ul_out.CN_dB, dl_out.CN_dB)
					worst_link = 'UL' if ul_out.CN_dB < dl_out.CN_dB else 'DL'
					worst_cn = min(ul_out.CN_dB, dl_out.CN_dB)
					e2e_status = self._assess_cn(worst_cn)[0] if worst_cn is not None else 'CRÍTICO'
					
					row.update({
						'e2e_latency_total_ms_one_way': round(ul_out.latency_ms + dl_out.latency_ms, 3),
						'e2e_latency_total_rtt_ms': round((ul_out.latency_ms + dl_out.latency_ms) * 2, 3),
						'e2e_cn_total_db': round(combined.get("CN_tot_dB", worst_cn), 2),
						'e2e_cinr_total_db': round(combined.get("CINR_tot_dB", worst_cn), 2),
						'e2e_worst_link': worst_link,
						'e2e_status': e2e_status,
					})
			
			# === POTENCIA Y BACK-OFF ===
			row.update({
				'eirp_sat_dbw': round(self.power_metrics.get('eirp_sat', float('nan')), 2),
				'input_backoff_db': round(self.power_metrics.get('input_bo', float('nan')), 2),
				'output_backoff_db': round(self.power_metrics.get('output_bo', float('nan')), 2),
				'eirp_dbw': round(self.core.eirp_dbw, 2),
			})
			
			# === RUIDO Y RENDIMIENTO ===
			if hasattr(self, 'perf_metrics'):
				pm = self.perf_metrics
				row.update({
					'T_sys_K': pm.get('T_sys_K', 150.0),
					'N0_dBHz': pm.get('N0_dBHz', -206.8),
					'EbN0_dB': pm.get('EbN0_dB'),
					'EbN0_req_dB': pm.get('EbN0_req_dB'),
					'Eb_margin_dB': pm.get('Eb_margin_dB'),
				})
			
			# Estado MODCOD
			if hasattr(self, 'modcod_status'):
				row['modcod_status'] = self.modcod_status
			
			# === GEOMETRÍA ORBITAL ===
			# Calcular valores orbitales reales
			orbital_radius_km = self.orbit_r_km if hasattr(self, 'orbit_r_km') else 42164  # GEO default
			orbital_velocity_ms = math.sqrt(MU_EARTH / (orbital_radius_km * 1000))  # m/s
			orbital_velocity_kms = orbital_velocity_ms / 1000  # km/s
			angular_velocity_rad_s = orbital_velocity_ms / (orbital_radius_km * 1000)  # rad/s
			angular_velocity_deg_s = angular_velocity_rad_s * 180 / math.pi  # deg/s
			orbital_period_s = 2 * math.pi * orbital_radius_km * 1000 / orbital_velocity_ms  # segundos
			orbital_period_min = orbital_period_s / 60  # minutos
			
			# Rate of change of distance (aproximación)
			range_rate_kms = 0  # Default, se puede calcular con posiciones anteriores
			
			row.update({
				'orbit_angle_deg': round(self.orbit_angle_deg, 2),
				'orbital_radius_km': round(orbital_radius_km, 2),
				'orbital_velocity_kms': round(orbital_velocity_kms, 3),
				'angular_velocity_degs': round(angular_velocity_deg_s, 6),
				'range_rate_kms': round(range_rate_kms, 3),
				'orbital_period_min': round(orbital_period_min, 2),
			})
			
			# === DOPPLER ===
			# Calcular Doppler instantáneo y máximo teórico
			if visible and hasattr(self, 'current_slant_distance_m'):
				# Doppler instantáneo basado en velocidad radial
				satellite_velocity_ms = orbital_velocity_ms  # m/s
				# Aproximación: component radial de velocidad = sat_vel * sin(elevation_angle)
				radial_velocity_ms = satellite_velocity_ms * math.sin(math.radians(self.current_elevation_deg))
				
				# Doppler shift: fd = (v_radial / c) * f_carrier
				c_light = 299792458  # m/s
				f_carrier_hz = 20e9  # 20 GHz típico
				doppler_hz = (radial_velocity_ms / c_light) * f_carrier_hz
				doppler_khz = doppler_hz / 1000
				
				# Doppler máximo teórico (horizonte)
				doppler_max_hz = (satellite_velocity_ms / c_light) * f_carrier_hz
				doppler_max_khz = doppler_max_hz / 1000
			else:
				doppler_khz = float('nan')
				doppler_max_khz = float('nan')
			
			row.update({
				'doppler_instantaneous_khz': round(doppler_khz, 3) if not math.isnan(doppler_khz) else None,
				'doppler_max_theoretical_khz': round(doppler_max_khz, 3) if not math.isnan(doppler_max_khz) else None,
			})
			
			# === PÉRDIDAS ===
			row['loss_total_extra_db'] = round(self.link_metrics.get('loss_total_extra_db', 0), 2) if 'loss_total_extra_db' in self.link_metrics else None
			
			# Pérdidas individuales
			for k, v in self.core.losses.items():
				row[k] = round(v, 2) if not math.isnan(v) else None
			
			self.history.append(row)

	def _toggle_losses(self):
		self.show_losses = not self.show_losses
		if self.show_losses:
			self.toggle_losses_btn.config(text='Ocultar Pérdidas ▼')
			# Posicionar filas detalladas al final
			base_row = 0
			for child in self.metrics_panel.grid_slaves():
				base_row = max(base_row, int(child.grid_info().get('row',0)))
			start = base_row + 1
			ordered = [
				('RFL_feeder','Feeder RF [dB]'),
				('AML_misalignment','Desalineación Antena [dB]'),
				('AA_atmos','Atenuación Atmosférica [dB]'),
				('Rain_att','Atenuación Lluvia [dB]'),
				('PL_polarization','Desajuste Polarización [dB]'),
				('L_pointing','Pérdida Apuntamiento [dB]'),
				('L_impl','Pérdidas Implementación [dB]'),
			]
			self._loss_title_labels = {}
			for i,(k,disp) in enumerate(ordered):
				# Crear title label y grid value label
				title = ttk.Label(self.metrics_panel, text=disp+':', font=('Segoe UI',10,'bold'), anchor='w')
				title.grid(row=start+i, column=0, sticky='w', padx=(2,4), pady=1)
				self._loss_title_labels[k] = title
				if k in self.loss_detail_label_map:
					self.loss_detail_label_map[k].grid(row=start+i, column=1, sticky='e', padx=(4,6), pady=1)
		else:
			self.toggle_losses_btn.config(text='Mostrar Pérdidas ▶')
			# Retirar labels
			if hasattr(self, '_loss_title_labels'):
				for k,lbl in self._loss_title_labels.items():
					if lbl.winfo_manager() == 'grid':
						lbl.grid_forget()
			for k,lbl in self.loss_detail_label_map.items():
				if lbl.winfo_manager() == 'grid':
					lbl.grid_forget()


	def _on_mousewheel(self, event):
		# Normalizar delta (Windows suele ser múltiplos de 120)
		if hasattr(self, 'metrics_canvas'):
			delta = int(-1*(event.delta/120))
			self.metrics_canvas.yview_scroll(delta, 'units')
	
	def _create_uplink_tab(self):
		"""Crea la pestaña de Uplink."""
		ul_frame = ttk.Frame(self.notebook)
		self.notebook.add(ul_frame, text='Uplink')
		
		# Variables específicas de UL
		self.ul_freq_var = tk.DoubleVar(value=self.link_state['UL'].f_Hz / 1e9)
		self.ul_bw_var = tk.DoubleVar(value=self.link_state['UL'].B_Hz / 1e6) 
		self.ul_eirp_var = tk.DoubleVar(value=self.link_state['UL'].EIRP_dBW)
		self.ul_gt_var = tk.DoubleVar(value=self.link_state['UL'].GT_dBK)
		self.ul_losses_var = tk.DoubleVar(value=self.link_state['UL'].L_extra_dB)
		
		# Conectar eventos de cambio
		for var in [self.ul_freq_var, self.ul_bw_var, self.ul_eirp_var, self.ul_gt_var, self.ul_losses_var]:
			var.trace('w', lambda *args: self.update_metrics())
		
		# Widgets
		ttk.Label(ul_frame, text='Frecuencia UL (GHz):').grid(row=0, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(ul_frame, textvariable=self.ul_freq_var, width=8).grid(row=0, column=1, padx=2, pady=1)
		
		ttk.Label(ul_frame, text='BW UL (MHz):').grid(row=1, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(ul_frame, textvariable=self.ul_bw_var, width=8).grid(row=1, column=1, padx=2, pady=1)
		
		ttk.Label(ul_frame, text='EIRP UL (dBW):').grid(row=2, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(ul_frame, textvariable=self.ul_eirp_var, width=8).grid(row=2, column=1, padx=2, pady=1)
		
		ttk.Label(ul_frame, text='G/T UL (dB/K):').grid(row=3, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(ul_frame, textvariable=self.ul_gt_var, width=8).grid(row=3, column=1, padx=2, pady=1)
		
		ttk.Label(ul_frame, text='Pérdidas Extra (dB):').grid(row=4, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(ul_frame, textvariable=self.ul_losses_var, width=8).grid(row=4, column=1, padx=2, pady=1)
		
		# Métricas de salida (solo lectura)
		self.ul_fspl_label = ttk.Label(ul_frame, text='FSPL: —', foreground='blue')
		self.ul_fspl_label.grid(row=5, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.ul_cn0_label = ttk.Label(ul_frame, text='C/N0: —', foreground='blue')
		self.ul_cn0_label.grid(row=6, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.ul_cn_label = ttk.Label(ul_frame, text='C/N: —', foreground='blue')
		self.ul_cn_label.grid(row=7, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.ul_lat_label = ttk.Label(ul_frame, text='Latencia: —', foreground='blue')
		self.ul_lat_label.grid(row=8, column=0, columnspan=2, sticky='w', padx=2, pady=1)
	
	def _create_downlink_tab(self):
		"""Crea la pestaña de Downlink."""
		dl_frame = ttk.Frame(self.notebook)
		self.notebook.add(dl_frame, text='Downlink')
		
		# Variables específicas de DL
		self.dl_freq_var = tk.DoubleVar(value=self.link_state['DL'].f_Hz / 1e9)
		self.dl_bw_var = tk.DoubleVar(value=self.link_state['DL'].B_Hz / 1e6)
		self.dl_eirp_var = tk.DoubleVar(value=self.link_state['DL'].EIRP_dBW)
		self.dl_gt_var = tk.DoubleVar(value=self.link_state['DL'].GT_dBK)
		self.dl_losses_var = tk.DoubleVar(value=self.link_state['DL'].L_extra_dB)
		
		# Conectar eventos de cambio
		for var in [self.dl_freq_var, self.dl_bw_var, self.dl_eirp_var, self.dl_gt_var, self.dl_losses_var]:
			var.trace('w', lambda *args: self.update_metrics())
		
		# Widgets
		ttk.Label(dl_frame, text='Frecuencia DL (GHz):').grid(row=0, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(dl_frame, textvariable=self.dl_freq_var, width=8).grid(row=0, column=1, padx=2, pady=1)
		
		ttk.Label(dl_frame, text='BW DL (MHz):').grid(row=1, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(dl_frame, textvariable=self.dl_bw_var, width=8).grid(row=1, column=1, padx=2, pady=1)
		
		ttk.Label(dl_frame, text='EIRP DL (dBW):').grid(row=2, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(dl_frame, textvariable=self.dl_eirp_var, width=8).grid(row=2, column=1, padx=2, pady=1)
		
		ttk.Label(dl_frame, text='G/T DL (dB/K):').grid(row=3, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(dl_frame, textvariable=self.dl_gt_var, width=8).grid(row=3, column=1, padx=2, pady=1)
		
		ttk.Label(dl_frame, text='Pérdidas Extra (dB):').grid(row=4, column=0, sticky='w', padx=2, pady=1)
		ttk.Entry(dl_frame, textvariable=self.dl_losses_var, width=8).grid(row=4, column=1, padx=2, pady=1)
		
		# Métricas de salida (solo lectura)
		self.dl_fspl_label = ttk.Label(dl_frame, text='FSPL: —', foreground='blue')
		self.dl_fspl_label.grid(row=5, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.dl_cn0_label = ttk.Label(dl_frame, text='C/N0: —', foreground='blue')
		self.dl_cn0_label.grid(row=6, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.dl_cn_label = ttk.Label(dl_frame, text='C/N: —', foreground='blue')
		self.dl_cn_label.grid(row=7, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.dl_lat_label = ttk.Label(dl_frame, text='Latencia: —', foreground='blue')
		self.dl_lat_label.grid(row=8, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		# Botones de copia
		btn_frame = ttk.Frame(dl_frame)
		btn_frame.grid(row=9, column=0, columnspan=2, pady=5)
		ttk.Button(btn_frame, text='← Copiar UL→DL', command=self._copy_ul_to_dl).pack(side='left', padx=2)
		ttk.Button(btn_frame, text='Copiar DL→UL →', command=self._copy_dl_to_ul).pack(side='left', padx=2)
	
	def _create_endtoend_tab(self):
		"""Crea la pestaña End-to-End."""
		e2e_frame = ttk.Frame(self.notebook)
		self.notebook.add(e2e_frame, text='End-to-End')
		
		# Métricas combinadas
		self.e2e_cn_ul_label = ttk.Label(e2e_frame, text='C/N (UL): —', foreground='green')
		self.e2e_cn_ul_label.grid(row=0, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.e2e_cn_dl_label = ttk.Label(e2e_frame, text='C/N (DL): —', foreground='green')
		self.e2e_cn_dl_label.grid(row=1, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.e2e_nc_tot_label = ttk.Label(e2e_frame, text='(N/C) Total: —', foreground='red')
		self.e2e_nc_tot_label.grid(row=2, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.e2e_cn_tot_label = ttk.Label(e2e_frame, text='C/N Total: —', foreground='red')
		self.e2e_cn_tot_label.grid(row=3, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.e2e_cinr_tot_label = ttk.Label(e2e_frame, text='CINR Total: —', foreground='red')
		self.e2e_cinr_tot_label.grid(row=4, column=0, columnspan=2, sticky='w', padx=2, pady=1)
		
		self.e2e_status_label = ttk.Label(e2e_frame, text='Estado: —', font=('Segoe UI', 10, 'bold'))
		self.e2e_status_label.grid(row=5, column=0, columnspan=2, sticky='w', padx=2, pady=5)
		
		# Selector de ancho de banda de referencia
		ttk.Label(e2e_frame, text='BW Referencia:').grid(row=6, column=0, sticky='w', padx=2, pady=1)
		self.bw_ref_var = tk.StringVar(value='DL')
		self.bw_ref_var.trace('w', lambda *args: self.update_metrics())
		ttk.Combobox(e2e_frame, textvariable=self.bw_ref_var, values=['UL', 'DL'], state='readonly', width=6).grid(row=6, column=1, padx=2, pady=1)
	
	def _on_tab_changed(self, event):
		"""Maneja el cambio de pestaña."""
		selected_tab = event.widget.tab('current')['text']
		if selected_tab == 'Uplink':
			self.current_link_sense = 'UL'
		elif selected_tab == 'Downlink':
			self.current_link_sense = 'DL'
		elif selected_tab == 'End-to-End':
			self.current_link_sense = 'E2E'
		self._sync_main_params_with_active_tab()
		self.update_metrics()

	def _sync_main_params_with_active_tab(self):
		"""Sincroniza parámetros principales con el tab activo UL/DL/E2E."""
		if self.current_link_sense == 'UL':
			self.eirp_var.set(self.ul_eirp_var.get())
			self.gt_var.set(self.ul_gt_var.get())
			self.bw_var.set(self.ul_bw_var.get())
		elif self.current_link_sense == 'DL':
			self.eirp_var.set(self.dl_eirp_var.get())
			self.gt_var.set(self.dl_gt_var.get())
			self.bw_var.set(self.dl_bw_var.get())
		elif self.current_link_sense == 'E2E':
			# Para End-to-End, usar los parámetros de referencia seleccionados
			ref_link = self.bw_ref_var.get()  # 'UL' o 'DL'
			if ref_link == 'UL':
				self.eirp_var.set(self.ul_eirp_var.get())
				self.gt_var.set(self.ul_gt_var.get())
				self.bw_var.set(self.ul_bw_var.get())
			else:  # DL
				self.eirp_var.set(self.dl_eirp_var.get())
				self.gt_var.set(self.dl_gt_var.get())
				self.bw_var.set(self.dl_bw_var.get())

	def _get_active_cn0_dbhz(self):
		"""Devuelve C/N0 del tab activo."""
		if hasattr(self, 'current_link_sense') and hasattr(self, 'link_out'):
			if self.current_link_sense == 'UL':
				ul_out = self.link_out.get('UL')
				return ul_out.CN0_dBHz if ul_out and ul_out.visible else float('nan')
			elif self.current_link_sense == 'DL':
				dl_out = self.link_out.get('DL')
				return dl_out.CN0_dBHz if dl_out and dl_out.visible else float('nan')
			elif self.current_link_sense == 'E2E':
				# Para End-to-End, usar C/N0 del enlace de referencia
				ref_link = self.bw_ref_var.get() if hasattr(self, 'bw_ref_var') else 'DL'
				ref_out = self.link_out.get(ref_link)
				return ref_out.CN0_dBHz if ref_out and ref_out.visible else float('nan')
		# Fallback al valor tradicional
		return self.link_metrics.get('cn0_dbhz', float('nan'))

	def _get_active_cn_db(self):
		"""Devuelve C/N del tab activo."""
		if hasattr(self, 'current_link_sense') and hasattr(self, 'link_out'):
			if self.current_link_sense == 'UL':
				ul_out = self.link_out.get('UL')
				return ul_out.CN_dB if ul_out and ul_out.visible else float('nan')
			elif self.current_link_sense == 'DL':
				dl_out = self.link_out.get('DL')
				return dl_out.CN_dB if dl_out and dl_out.visible else float('nan')
			elif self.current_link_sense == 'E2E':
				# Para End-to-End, mostrar C/N total combinado
				ul_out = self.link_out.get('UL')
				dl_out = self.link_out.get('DL')
				if ul_out and dl_out and ul_out.visible and dl_out.visible:
					combined = combine_end_to_end(ul_out.CN_dB, dl_out.CN_dB)
					return combined["CN_tot_dB"]
				return float('nan')
		# Fallback al valor tradicional
		return self.link_metrics.get('cn_db', float('nan'))

	def _get_active_fspl_db(self):
		"""Devuelve FSPL del tab activo."""
		if hasattr(self, 'current_link_sense') and hasattr(self, 'link_out'):
			if self.current_link_sense == 'UL':
				ul_out = self.link_out.get('UL')
				return ul_out.FSPL_dB if ul_out and ul_out.visible else float('nan')
			elif self.current_link_sense == 'DL':
				dl_out = self.link_out.get('DL')
				return dl_out.FSPL_dB if dl_out and dl_out.visible else float('nan')
			elif self.current_link_sense == 'E2E':
				# Para End-to-End, usar FSPL del enlace de referencia
				ref_link = self.bw_ref_var.get() if hasattr(self, 'bw_ref_var') else 'DL'
				ref_out = self.link_out.get(ref_link)
				return ref_out.FSPL_dB if ref_out and ref_out.visible else float('nan')
		# Fallback al valor tradicional
		return self.link_metrics.get('fspl_db', float('nan'))

	def _get_active_ebn0_db(self):
		"""Devuelve Eb/N0 del tab activo para evaluación MODCOD."""
		if hasattr(self, 'current_link_sense') and hasattr(self, 'link_out'):
			if self.current_link_sense == 'UL':
				# Usar C/N UL directamente (más simple y directo)
				ul_out = self.link_out.get('UL')
				if ul_out and ul_out.visible:
					return ul_out.CN_dB
				return float('nan')
			elif self.current_link_sense == 'DL':
				# Usar C/N DL directamente
				dl_out = self.link_out.get('DL')
				if dl_out and dl_out.visible:
					return dl_out.CN_dB
				return float('nan')
			elif self.current_link_sense == 'E2E':
				# Para End-to-End, usar el peor caso (menor C/N)
				ul_out = self.link_out.get('UL')
				dl_out = self.link_out.get('DL')
				if ul_out and dl_out and ul_out.visible and dl_out.visible:
					return min(ul_out.CN_dB, dl_out.CN_dB)
				elif ul_out and ul_out.visible:
					return ul_out.CN_dB
				elif dl_out and dl_out.visible:
					return dl_out.CN_dB
				return float('nan')
		# Fallback al valor tradicional
		return self.perf_metrics.get('EbN0_dB', float('nan'))

	
	def _copy_ul_to_dl(self):
		"""Copia parámetros de UL a DL."""
		self.dl_freq_var.set(self.ul_freq_var.get())
		self.dl_bw_var.set(self.ul_bw_var.get())
		self.dl_eirp_var.set(self.ul_eirp_var.get())
		self.dl_gt_var.set(self.ul_gt_var.get())
		self.dl_losses_var.set(self.ul_losses_var.get())
		self.update_metrics()
	
	def _copy_dl_to_ul(self):
		"""Copia parámetros de DL a UL."""
		self.ul_freq_var.set(self.dl_freq_var.get())
		self.ul_bw_var.set(self.dl_bw_var.get())
		self.ul_eirp_var.set(self.dl_eirp_var.get())
		self.ul_gt_var.set(self.dl_gt_var.get())
		self.ul_losses_var.set(self.dl_losses_var.get())
		self.update_metrics()
	
	def _compute_link_outputs_local(self, inputs: LinkInputs, d_km: float) -> LinkOutputs:
		"""Calcula las métricas de salida para un sentido de enlace (versión local)."""
		# Verificar visibilidad basada en elevación y distancia válida
		elevation_deg = getattr(self, 'current_elevation_deg', 0.0)
		# Usar la elevación real mostrada en la GUI en lugar de la calculada internamente
		if hasattr(self, 'current_visible') and self.current_visible:
			visible = True
		else:
			visible = d_km > 0 and elevation_deg > 0.1  # usar 0.1 en lugar de valores muy pequeños
		
		if not visible:
			return LinkOutputs(
				FSPL_dB=float('nan'),
				CN0_dBHz=float('nan'), 
				CN_dB=float('nan'),
				visible=False,
				latency_ms=float('nan')
			)
		
		# Cálculos usando las funciones puras
		fspl = fspl_dB(inputs.f_Hz, d_km * 1000)  # convertir km a m
		cn0 = cn0_dBHz(inputs.EIRP_dBW, inputs.GT_dBK, fspl, inputs.L_extra_dB)
		cn = cn_dB(cn0, inputs.B_Hz)
		latency = (d_km * 1000 / SPEED_OF_LIGHT) * 1000  # ms
		
		return LinkOutputs(
			FSPL_dB=fspl,
			CN0_dBHz=cn0,
			CN_dB=cn,
			visible=True,
			latency_ms=latency
		)

	def _update_link_params(self):
		"""Actualiza los parámetros de los enlaces UL/DL desde las variables de la GUI."""		
		if hasattr(self, 'ul_freq_var'):
			# Actualizar estado UL
			self.link_state['UL'].f_Hz = self.ul_freq_var.get() * 1e9
			self.link_state['UL'].B_Hz = self.ul_bw_var.get() * 1e6
			self.link_state['UL'].EIRP_dBW = self.ul_eirp_var.get()
			self.link_state['UL'].GT_dBK = self.ul_gt_var.get()
			self.link_state['UL'].L_extra_dB = self.ul_losses_var.get()
		
		if hasattr(self, 'dl_freq_var'):
			# Actualizar estado DL
			self.link_state['DL'].f_Hz = self.dl_freq_var.get() * 1e9
			self.link_state['DL'].B_Hz = self.dl_bw_var.get() * 1e6
			self.link_state['DL'].EIRP_dBW = self.dl_eirp_var.get()
			self.link_state['DL'].GT_dBK = self.dl_gt_var.get()
			self.link_state['DL'].L_extra_dB = self.dl_losses_var.get()
		
		# Calcular métricas para ambos enlaces usando el método local
		d_km = getattr(self, 'current_slant_distance_m', 600000.0) / 1000.0
		
		for link in ['UL', 'DL']:
			inputs = self.link_state[link]
			self.link_out[link] = self._compute_link_outputs_local(inputs, d_km)
	
	def _update_link_gui(self):
		"""Actualiza las etiquetas de la GUI con las métricas calculadas."""
		import math
		
		if hasattr(self, 'ul_fspl_label'):
			# Actualizar UL
			ul_out = self.link_out['UL']
			if ul_out.visible and not math.isnan(ul_out.FSPL_dB):
				self.ul_fspl_label.config(text=f'FSPL: {ul_out.FSPL_dB:.1f} dB')
				self.ul_cn0_label.config(text=f'C/N0: {ul_out.CN0_dBHz:.1f} dB-Hz')
				self.ul_cn_label.config(text=f'C/N: {ul_out.CN_dB:.1f} dB')
				self.ul_lat_label.config(text=f'Latencia: {ul_out.latency_ms:.2f} ms')
			else:
				self.ul_fspl_label.config(text='FSPL: —')
				self.ul_cn0_label.config(text='C/N0: —')
				self.ul_cn_label.config(text='C/N: —')
				self.ul_lat_label.config(text='Latencia: —')
		
		if hasattr(self, 'dl_fspl_label'):
			# Actualizar DL
			dl_out = self.link_out['DL']
			if dl_out.visible and not math.isnan(dl_out.FSPL_dB):
				self.dl_fspl_label.config(text=f'FSPL: {dl_out.FSPL_dB:.1f} dB')
				self.dl_cn0_label.config(text=f'C/N0: {dl_out.CN0_dBHz:.1f} dB-Hz')
				self.dl_cn_label.config(text=f'C/N: {dl_out.CN_dB:.1f} dB')
				self.dl_lat_label.config(text=f'Latencia: {dl_out.latency_ms:.2f} ms')
			else:
				self.dl_fspl_label.config(text='FSPL: —')
				self.dl_cn0_label.config(text='C/N0: —')
				self.dl_cn_label.config(text='C/N: —')
				self.dl_lat_label.config(text='Latencia: —')
		
		if hasattr(self, 'e2e_cn_ul_label'):
			# Actualizar End-to-End
			ul_out = self.link_out['UL']
			dl_out = self.link_out['DL']
			
			# Solo calcular si ambos enlaces son válidos
			if (ul_out.visible and dl_out.visible and 
				not math.isnan(ul_out.CN_dB) and not math.isnan(dl_out.CN_dB)):
				
				# Combinar End-to-End
				combined = combine_end_to_end(ul_out.CN_dB, dl_out.CN_dB)
				
				self.e2e_cn_ul_label.config(text=f'C/N (UL): {ul_out.CN_dB:.1f} dB')
				self.e2e_cn_dl_label.config(text=f'C/N (DL): {dl_out.CN_dB:.1f} dB')
				self.e2e_nc_tot_label.config(text=f'(N/C) Total: {combined["NC_tot_dB"]:.2f} dB')
				self.e2e_cn_tot_label.config(text=f'C/N Total: {combined["CN_tot_dB"]:.1f} dB')
				self.e2e_cinr_tot_label.config(text=f'CINR Total: {combined["CINR_tot_dB"]:.1f} dB')
				
				# Estado del enlace
				status = 'NOMINAL' if combined["CN_tot_dB"] > 10.0 else 'MARGINAL' if combined["CN_tot_dB"] > 5.0 else 'CRÍTICO'
				color = 'green' if status == 'NOMINAL' else 'orange' if status == 'MARGINAL' else 'red'
				self.e2e_status_label.config(text=f'Estado: {status}', foreground=color)
			else:
				# No hay datos válidos - satélite no visible
				self.e2e_cn_ul_label.config(text='C/N (UL): —')
				self.e2e_cn_dl_label.config(text='C/N (DL): —')
				self.e2e_nc_tot_label.config(text='(N/C) Total: —')
				self.e2e_cn_tot_label.config(text='C/N Total: —')
				self.e2e_cinr_tot_label.config(text='CINR Total: —')
				self.e2e_status_label.config(text='Estado: No visible', foreground='gray')


# ------------------------------------------------------------- #
# Punto de entrada                                              #
# ------------------------------------------------------------- #
def main():
	loader = ParameterLoader()
	core = JammerSimulatorCore(loader)
	if tk is None:
		print("Tkinter no disponible en este entorno.")
		return
	root = tk.Tk()
	SimulatorGUI(root, core)
	root.mainloop()


if __name__ == '__main__':
	main()

